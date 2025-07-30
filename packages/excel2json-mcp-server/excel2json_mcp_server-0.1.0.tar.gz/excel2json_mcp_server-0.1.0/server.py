import os
import re
import json
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook
from mcp.server.fastmcp import FastMCP

# 创建 MCP 服务器实例
mcp = FastMCP("Excel to JSON Converter")

def _clean_column_names(columns):
    """安全处理列名，保留中文和其他非ASCII字符，处理空列名和pandas自动生成的Unnamed列"""
    cleaned = []
    for i, col in enumerate(columns):
        if pd.isna(col):
            col_str = f"column_{i + 1}"
        else:
            col_str = str(col).strip()
            if not col_str or col_str.isspace() or col_str.startswith("Unnamed:"):
                col_str = f"column_{i + 1}"
            else:
                col_str = re.sub(r'[^\w\s.\-\u4e00-\u9fff]', '', col_str)
                col_str = re.sub(r'\s+', '_', col_str)
        cleaned.append(col_str)
    return cleaned


def _column_letter_to_index(letter):
    """将 Excel 列字母转换为 1-based 索引（A=1, B=2, ..., AA=27, AB=28, ...）"""
    index = 0
    for char in letter.upper():
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index


def _get_merged_cells_info(ws):
    """获取指定工作表的合并单元格范围，并返回详细信息"""
    merged_cells_info = []
    for rng in ws.merged_cells.ranges:
        range_str = str(rng)
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str)
        if match:
            start_col_letter, start_row, end_col_letter, end_row = match.groups()
            start_row = int(start_row)
            end_row = int(end_row)
            start_col_idx = _column_letter_to_index(start_col_letter)
            end_col_idx = _column_letter_to_index(end_col_letter)
            value = ws.cell(start_row, start_col_idx).value
            merged_cells_info.append({
                "start_row": start_row,
                "end_row": end_row,
                "start_col": start_col_idx,
                "end_col": end_col_idx,
                "value": value
            })
    return merged_cells_info


def _make_names_unique(names: list[str]) -> list[str]:
    """确保列表中的所有名称唯一，如果需要会追加计数器"""
    counts = {}
    unique_names = []
    for name in names:
        original_name = name
        current_name = name
        if name in counts:
            counts[original_name] += 1
            current_name = f"{original_name}_{counts[original_name]}"
        else:
            counts[original_name] = 1
        unique_names.append(current_name)
    return unique_names


def _parse_single_header(ws, merged_cells_info: list) -> tuple[list[str], int]:
    actual_header_row_excel_num = 0
    max_col_from_header = 0

    for r_num in range(1, ws.max_row + 1):
        is_current_row_empty = True
        current_row_max_col = 0

        row_has_cells = False
        for row_tuple in ws.iter_rows(min_row=r_num, max_row=r_num, values_only=False):
            row_has_cells = True
            for cell in row_tuple:
                if cell.value is not None:
                    is_current_row_empty = False
                    current_row_max_col = max(current_row_max_col, cell.column)
            break
        if not row_has_cells and r_num > ws.max_row:
            break

        if not is_current_row_empty:
            actual_header_row_excel_num = r_num
            max_col_from_header = current_row_max_col
            break

    if actual_header_row_excel_num == 0:
        return [], 0

    header_rows_to_skip = actual_header_row_excel_num
    num_cols_to_process = max_col_from_header

    raw_header_values: List[Optional[str]] = [None] * num_cols_to_process

    for c_idx in range(num_cols_to_process):
        current_col_num = c_idx + 1
        cell_value = ws.cell(row=actual_header_row_excel_num, column=current_col_num).value

        for merged_info in merged_cells_info:
            if merged_info["start_row"] == actual_header_row_excel_num and \
                    merged_info["end_row"] == actual_header_row_excel_num and \
                    merged_info["start_col"] <= current_col_num <= merged_info["end_col"]:
                cell_value = ws.cell(row=merged_info["start_row"], column=merged_info["start_col"]).value
                break

        if cell_value is not None and str(cell_value).strip():
            raw_header_values[c_idx] = str(cell_value).strip()
        else:
            raw_header_values[c_idx] = f"column_{current_col_num}"

    cleaned_names = [_clean_column_names([name])[0] for name in raw_header_values]
    final_column_names = _make_names_unique(cleaned_names)

    return final_column_names, header_rows_to_skip


def _normalize_and_validate_file_path(file_path: str) -> str:
    """规范化文件路径并验证其有效性"""
    try:
        normalized_path = os.path.abspath(os.path.normpath(file_path))
    except Exception as error:
        error_response = {
            "path_error": {
                "type": "InvalidFilePath",
                "message": f"Invalid file path: {str(error)}",
                "suggestion": "Please provide a valid file path."
            }
        }
        return json.dumps(error_response, indent=2, ensure_ascii=False)

    if not os.path.isfile(normalized_path):
        error_response = {
            "path_error": {
                "type": "FileNotFound",
                "message": f"The file '{normalized_path}' does not exist.",
                "suggestion": "Please check the file path and try again."
            }
        }
        return json.dumps(error_response, indent=2, ensure_ascii=False)

    return normalized_path


@mcp.tool()
def excel_to_json(
        file_path: str,
        sheet_name_param: str | int | list[str | int] | None = 0
) -> str:
    """
    将Excel文件的工作表转换为JSON格式。

    参数:
        file_path (str): Excel文件的完整路径，必须是.xlsx格式
        sheet_name_param (str | int | list | None): 要处理的工作表
            - int: 工作表索引（从0开始）
            - str: 工作表名称
            - list: 多个工作表名称或索引的列表
            - None: 处理所有工作表
            - 默认值: 0 (第一个工作表)

    返回:
        str: JSON格式的字符串，包含Excel数据

    使用方法:
        # 处理第一个工作表
        result = excel_to_json("/path/to/file.xlsx")

        # 处理指定名称的工作表
        result = excel_to_json("/path/to/file.xlsx", "Sheet1")

        # 处理多个工作表
        result = excel_to_json("/path/to/file.xlsx", ["Sheet1", "Sheet2"])

        # 处理所有工作表
        result = excel_to_json("/path/to/file.xlsx", None)
    """
    normalized_path = _normalize_and_validate_file_path(file_path)
    if normalized_path.startswith('{"path_error":'):
        return normalized_path

    file_ext = os.path.splitext(normalized_path)[1].lower()
    if file_ext != ".xlsx":
        error_response = {
            "suffix_error": {
                "type": "UnsupportedFileFormat",
                "message": f"The file format '{file_ext}' is not supported. "
                           f"This tool currently only processes .xlsx files. "
                           f"Please convert your file to the .xlsx format and try again.",
                "suggestion": "Convert the file to .xlsx format."
            }
        }
        return json.dumps(error_response)

    try:
        wb = load_workbook(normalized_path, data_only=True)
    except FileNotFoundError:
        raise
    except Exception as error:
        raise Exception(f"Could not load workbook: {error}")

    sheets_to_process_names = []
    if sheet_name_param is None:
        sheets_to_process_names = wb.sheetnames
    elif isinstance(sheet_name_param, (list, tuple)):
        for sn_item in sheet_name_param:
            if isinstance(sn_item, int):
                if 0 <= sn_item < len(wb.sheetnames):
                    sheets_to_process_names.append(wb.sheetnames[sn_item])
                else:
                    wb.close()
                    raise ValueError(f"Sheet index {sn_item} is out of range.")
            else:
                if sn_item in wb.sheetnames:
                    sheets_to_process_names.append(sn_item)
                else:
                    wb.close()
                    raise ValueError(f"Sheet name '{sn_item}' not found.")
    else:
        if isinstance(sheet_name_param, int):
            if 0 <= sheet_name_param < len(wb.sheetnames):
                sheets_to_process_names.append(wb.sheetnames[sheet_name_param])
            else:
                wb.close()
                raise ValueError(f"Sheet index {sheet_name_param} is out of range.")
        else:
            if sheet_name_param in wb.sheetnames:
                sheets_to_process_names.append(sheet_name_param)
            else:
                wb.close()
                raise ValueError(f"Sheet name '{sheet_name_param}' not found.")

    if not sheets_to_process_names:
        wb.close()
        return json.dumps({}, indent=2, ensure_ascii=False)

    result_data = {}
    for sheet_name_str in sheets_to_process_names:
        ws = wb[sheet_name_str]

        merged_cells_info = _get_merged_cells_info(ws)
        column_names, header_row_excel_num = _parse_single_header(ws, merged_cells_info)

        data_list_for_df = []
        excel_indices_for_df = []
        data_start_excel_num = header_row_excel_num + 1 if header_row_excel_num > 0 else 1

        for r_idx_offset, row_cells in enumerate(ws.iter_rows(min_row=data_start_excel_num)):
            actual_excel_row = data_start_excel_num + r_idx_offset

            is_row_entirely_empty = True
            current_row_original_values = []
            num_cols_in_current_row_from_ws = len(row_cells)
            num_cols_to_process_for_row = len(column_names) if column_names else num_cols_in_current_row_from_ws

            for c_check_idx in range(num_cols_in_current_row_from_ws):
                cell_obj = row_cells[c_check_idx]
                val_for_check = cell_obj.value
                current_row_original_values.append(val_for_check)

                if c_check_idx < num_cols_to_process_for_row and val_for_check is not None:
                    is_row_entirely_empty = False

            if is_row_entirely_empty and not data_list_for_df:
                continue

            current_row_data_dict = {}
            has_any_value_in_this_dict_row = False

            if column_names:
                for c_data_idx, col_name in enumerate(column_names):
                    cell_value_for_dict = None
                    if c_data_idx < len(current_row_original_values):
                        cell_value_for_dict = current_row_original_values[c_data_idx]

                    for merged_info in merged_cells_info:
                        if merged_info["start_row"] <= actual_excel_row <= merged_info["end_row"] and \
                                merged_info["start_col"] <= (c_data_idx + 1) <= merged_info["end_col"]:
                            cell_value_for_dict = ws.cell(row=merged_info["start_row"],
                                                          column=merged_info["start_col"]).value
                            break

                    current_row_data_dict[col_name] = cell_value_for_dict
                    if cell_value_for_dict is not None:
                        has_any_value_in_this_dict_row = True
            else:
                for c_data_idx, cell_val in enumerate(current_row_original_values):
                    final_cell_val = cell_val
                    for merged_info in merged_cells_info:
                        if merged_info["start_row"] <= actual_excel_row <= merged_info["end_row"] and \
                                merged_info["start_col"] <= (c_data_idx + 1) <= merged_info["end_col"]:
                            final_cell_val = ws.cell(row=merged_info["start_row"],
                                                     column=merged_info["start_col"]).value
                            break
                    current_row_data_dict[f"column_{c_data_idx + 1}"] = final_cell_val
                    if final_cell_val is not None:
                        has_any_value_in_this_dict_row = True

            should_add_row_to_df = False
            if has_any_value_in_this_dict_row:
                should_add_row_to_df = True
            elif data_list_for_df and not is_row_entirely_empty:
                should_add_row_to_df = True

            if should_add_row_to_df:
                data_list_for_df.append(current_row_data_dict)
                excel_indices_for_df.append(actual_excel_row)

        df_columns = column_names if column_names else None
        if not column_names and data_list_for_df:
            if data_list_for_df[0]:
                df_columns = list(data_list_for_df[0].keys())

        df = pd.DataFrame(data_list_for_df, columns=df_columns)

        index_name = "excel_row_number"

        if not df.empty and excel_indices_for_df:
            df.index = pd.Index(excel_indices_for_df, name=index_name)
        elif not df.empty:
            df.index = pd.RangeIndex(start=data_start_excel_num, stop=data_start_excel_num + len(df),
                                     name=index_name)

        if df.empty:
            df.index = pd.Index([], name=index_name)

        json_output_str = df.to_json(orient="table", index=True, force_ascii=False)
        sheet_json_data = json.loads(json_output_str)

        result_data[sheet_name_str] = sheet_json_data

    wb.close()
    return json.dumps(result_data, indent=2, ensure_ascii=False)

def main():
    mcp.run()

if __name__ == "__main__":
    main()
