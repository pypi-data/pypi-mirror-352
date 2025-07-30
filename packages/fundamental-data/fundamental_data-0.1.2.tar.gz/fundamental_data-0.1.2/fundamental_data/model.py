import pandas as pd
import numpy as np
import yfinance as yf
from collections import defaultdict
class StockFundamentals:
    def __init__(self, data, ticker):
        self.data = data
        self.ticker = ticker
        self.result = self._go()
        self.raw_data = data['facts']['us-gaap']
        
        # Assign tables as attributes
        self.QuarterlyTable = self.result['QuarterlyTable']
        self.DatedTable = self.result['DatedTable']
        self.MetricStats = self.result['MetricStats']
        self.CombinedTable = self._create_combined_table()
        self.VisualizationTable = self._create_visualization_table()
        
    def _go(self):
        # Initialize containers
        self.item_store = defaultdict(list)
        self.metric_stats = []
        self.all_ends = []
        
        # Process JSON data
        self._process_data()
        self.earliestDate = min(self.all_ends) if self.all_ends else pd.Timestamp.now() - pd.DateOffset(years=5)
        self.latestDate = max(self.all_ends) if self.all_ends else pd.Timestamp.now()
        
        # Create tables
        return {
            'QuarterlyTable': self._create_quarterly_table(),
            'DatedTable': self._create_dated_table(),
            'MetricStats': self._create_metric_stats()
        }
    
    def _process_data(self):
        gaap_facts = self.data['facts']['us-gaap']
        
        for key in gaap_facts.keys():
            fact = gaap_facts[key]
            if 'Deprecated' in str(fact.get('label', '')):
                continue
                
            units = fact['units']
            if len(units) != 1:
                continue
                
            unit_key, unit_items = next(iter(units.items()))
            self._process_items(key, unit_key, unit_items)

    def _process_items(self, metric, unit, items):
        valid_items = []
        for item in items:
            if 'frame' in item and 'Q' in item.get('frame', ''):
                end_date = pd.to_datetime(item['end'], format='%Y-%m-%d')
                frame = item['frame'].strip('I').strip('CY')
                
                self.item_store[metric].append({
                    'end': end_date,
                    'value': item.get('val', np.nan),
                    'frame': frame,
                    'unit': unit,
                    'form': item.get('form', '')
                })
                self.all_ends.append(end_date)
                
        if self.item_store[metric]:
            dates = [i['end'] for i in self.item_store[metric]]
            self.metric_stats.append({
                'metric': metric,
                'unit': unit,
                'first_date': min(dates),
                'last_date': max(dates),
                'entry_count': len(dates),
                'forms': ','.join(sorted({i['form'] for i in self.item_store[metric]}))
            })

    def _create_quarterly_table(self):
        quarterly_data = []
        for metric, items in self.item_store.items():
            for item in items:
                try:
                    quarterly_data.append({
                        'quarter': pd.Period(item['frame'], freq='Q'),
                        'metric': metric,
                        'value': item['value']
                    })
                except:
                    continue
                    
        df = pd.DataFrame(quarterly_data)
        return df.pivot_table(
            index='quarter',
            columns='metric',
            values='value',
            aggfunc='first'
        ).sort_index()

    def _create_dated_table(self):
        date_data = []
        for metric, items in self.item_store.items():
            for item in items:
                if '10-Q' in item['form']:
                    date_data.append({
                        'end_date': item['end'],
                        'metric': metric,
                        'value': item['value']
                    })
                    
        df = pd.DataFrame(date_data)
        return df.pivot_table(
            index='end_date',
            columns='metric',
            values='value',
            aggfunc='first'
        ).sort_index()

    def _create_metric_stats(self):
        # Calculate streak statistics for each metric
        for stats in self.metric_stats:
            metric = stats['metric']
            # Convert to ordinal for streak calculations
            quarters = sorted([pd.Period(item['frame'], freq='Q').ordinal 
                            for item in self.item_store[metric]])
            
            # Calculate streaks using ordinal values
            streaks = []
            current_streak = 1
            for i in range(1, len(quarters)):
                if quarters[i] == quarters[i-1] + 1:
                    current_streak += 1
                else:
                    streaks.append(current_streak)
                    current_streak = 1
            streaks.append(current_streak)
            
            stats['longest_streak'] = max(streaks) if streaks else 0
            stats['break_count'] = len(streaks) - 1

        df = pd.DataFrame(self.metric_stats)
        return df.set_index('metric').sort_values(
            ['entry_count', 'metric'], 
            ascending=[False, True]
        )[['unit', 'first_date', 'last_date', 'entry_count', 
           'forms', 'longest_streak', 'break_count']]

    def _create_visualization_table(self):
        # Get all quarters from min to max
        start = self.QuarterlyTable.index.min().to_timestamp()
        end = self.QuarterlyTable.index.max().to_timestamp()
        full_index = pd.period_range(start=start, end=end, freq='Q')
        
        # Create empty DataFrame with all quarters
        full_presence = pd.DataFrame(index=full_index)
        
        # Mark available data (1=available, 0=missing)
        for metric in self.QuarterlyTable.columns:
            available = self.QuarterlyTable[metric].notna().astype(int)
            full_presence[metric] = available.reindex(full_index, fill_value=0)
        
        # Add quarter info without duplicate columns
        full_presence = full_presence.reset_index()
        full_presence = full_presence.rename(columns={'index': 'quarter'})
        full_presence['year'] = full_presence['quarter'].dt.year
        full_presence['qtr'] = full_presence['quarter'].dt.quarter
        
        return full_presence

    def visualize_data_availability(self):
        """Create a scrollable visualization of data availability"""
        import plotly.express as px
        
        # Melt the data for visualization
        viz_df = self.VisualizationTable.melt(
            id_vars=['quarter', 'year', 'qtr'],
            var_name='metric',
            value_name='available'
        )
        
        # Convert quarter to string for better display
        viz_df['quarter_str'] = viz_df['quarter'].dt.strftime('%Y-Q%q')
        
        fig = px.scatter(
            viz_df,
            x='quarter_str',
            y='metric',
            color='available',
            color_continuous_scale=['red', 'green'],
            title=f"Data Availability for {self.ticker}",
            labels={'quarter_str': 'Quarter', 'metric': 'Financial Metric'}
        )
        
        fig.update_layout(
            height=800,
            xaxis_rangeslider_visible=True,
            coloraxis_showscale=False,
            xaxis={'tickangle': 45}
        )
        
        return fig.show()

    def _create_combined_table(self):
        # Get price data without MultiIndex
        price_data = yf.download(
            self.ticker,
            start=self.earliestDate,
            end=self.latestDate
        )
        
        # Flatten the columns
        price_data.columns = [col[0].lower() for col in price_data.columns.values]
        
        # Merge with DatedTable
        return pd.merge_asof(
            self.DatedTable.sort_index(),
            price_data[['close']].rename(columns={'close': 'price'}),
            left_index=True,
            right_index=True,
            direction='nearest'
        )
    
    def saveVisualizationTable(self, xlsx_path):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from pandas import ExcelWriter

        # Save the DataFrame to an Excel file first (without formatting)
        self.VisualizationTable.to_excel(xlsx_path, index=False, sheet_name='Availability')

        # Open the workbook and select the sheet
        wb = load_workbook(xlsx_path)
        ws = wb['Availability']

        # Define fills
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Determine the range of data cells (excluding header row)
        for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                # Skip year/qtr/quarter columns
                if isinstance(cell.value, int) and cell.column > 3:
                    if cell.value == 1:
                        cell.fill = green_fill
                    elif cell.value == 0:
                        cell.fill = red_fill

        # Save the workbook with styles
        wb.save(xlsx_path)