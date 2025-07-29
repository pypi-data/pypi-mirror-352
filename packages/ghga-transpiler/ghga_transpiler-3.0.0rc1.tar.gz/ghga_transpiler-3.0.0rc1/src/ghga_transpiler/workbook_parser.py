# Copyright 2021 - 2025 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"Module containing logic to parse a GHGA workbook"

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ConfigDict, Field

from .config import WorkbookConfig, WorksheetSettings
from .models import GHGAWorkbook, GHGAWorksheet


class WorksheetParser(BaseModel):
    """Group worksheet parser functions."""

    config: WorksheetSettings

    def _header(self, worksheet: Worksheet):
        """Return a list of column names of a worksheet."""
        return [
            cell.value
            for row in worksheet.iter_rows(
                self.config.settings.header_row,
                self.config.settings.header_row,
                self.config.settings.start_column,
                self.config.settings.end_column,
            )
            for cell in row
        ]

    def _rows(self, worksheet: Worksheet) -> list:
        """Create a list of non-empty rows of a worksheet."""
        return [
            row
            for row in worksheet.iter_rows(
                self.config.settings.start_row,
                worksheet.max_row,
                self.config.settings.start_column,
                self.config.settings.end_column,
                values_only=True,
            )
            if not all(cell is None for cell in row)
        ]

    def _content(self, worksheet: Worksheet) -> list[dict]:
        """Compute and return the content of the worksheet, rows as worksheet row values and
        column names as keys
        """
        return [
            {
                key: value
                for key, value in zip(self._header(worksheet), row, strict=True)
                if value is not None and value != ""
            }
            for row in self._rows(worksheet)
        ]

    def _transformed_content(self, worksheet: Worksheet) -> list:
        """Processes each row of the provided worksheet, applying transformations to
        specific fields as defined in the configuration and returns the worksheet content.
        """
        transformed_data = []
        for row in self._content(worksheet):
            transformed_row = {}
            for key, value in row.items():
                transformations = self.config.get_transformations()
                if transformations and key in transformations:
                    transformed_row[key] = transformations[key](value)
                else:
                    transformed_row[key] = value
            transformed_data.append(transformed_row)
        return transformed_data


class GHGAWorksheetParser(WorksheetParser):
    """Extend WorksheetParser with GHGA worksheet specific parsers."""

    def parse(self, worksheet: Worksheet):
        """Render a worksheet into GHGAWorksheet model"""
        return GHGAWorksheet.model_validate(
            {"worksheet": {self.config.settings.name: self._parse(worksheet)}}
        )

    def _parse(self, worksheet: Worksheet) -> dict[str, dict]:
        """Parse a worksheet row by row into a dictionary of row-primary-keys as keys and
        a dictionary of content and relations as the values.
        """
        worksheet_data = self._transformed_content(worksheet)
        return {
            row[self.config.settings.primary_key]: {
                "content": self._relation_free_content(row),
                "relations": self._relations(row),
            }
            for row in worksheet_data
        }

    def _relations(self, row: dict) -> dict:
        """Get relations to a dictionary that contains relation name as key and the
        resource that is in the relation as the value
        """
        relations = self.config.get_relations()
        return {
            relation.name: {
                "targetClass": relation.target_class,
                "targetResources": row[relation.name],
            }
            for relation in relations
            if relation.name in row
        }

    def _relation_free_content(self, row: dict) -> dict:
        """Clean up the content data from the relation, i.e., remove the key value pairs
        belonging to a relation from the row content.
        """
        relations = self._relations(row)
        relations[self.config.settings.primary_key] = None
        return {key: value for key, value in row.items() if key not in relations}


class GHGAWorkbookParser(BaseModel):
    """Parser class for converting a workbook into a GHGAWorkbook."""

    model_config = ConfigDict(arbitrary_types_allowed=True)
    config: WorkbookConfig = Field(
        ...,
        description="Configuration for processing the workbook, including worksheet"
        + " settings and column transformations.",
    )

    workbook: Workbook = Field(
        ...,
        description="Path to the Excel workbook file (.xlsx) that will be parsed."
        + " This file contains the data to be transformed and processed.",
    )

    exclude: list = Field(
        default=[
            "__transpiler_protocol",
            "__sheet_meta",
            "__column_meta",
        ],
        description="List of sheet names to exclude from processing."
        + " These are typically meta sheets that contain configuration data or metadata"
        + " rather than actual worksheet data. Default value corresponds to the GHGA"
        + " standard configuration sheets on the GHGA submission workbook.",
    )

    def parse(self) -> GHGAWorkbook:
        """Converts the given workbook into a GHGAWorkbook instance.

        This method iterates through the sheets of the provided workbook, excluding
        any meta sheets (i.e., '__transpiler_protocol', '__sheet_meta', '__column_meta').
        """
        return GHGAWorkbook.model_validate(
            {
                "workbook": tuple(
                    GHGAWorksheetParser(config=self.config.worksheets[name]).parse(
                        self.workbook[name]
                    )
                    for name in self.workbook.sheetnames
                    if name not in self.exclude
                )
            }
        )
