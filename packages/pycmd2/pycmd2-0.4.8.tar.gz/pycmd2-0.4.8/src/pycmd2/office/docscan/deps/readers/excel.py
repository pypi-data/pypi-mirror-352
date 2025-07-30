import pathlib
import typing

import openpyxl
import xlrd

from pycmd2.office.docscan.deps.readers.base import BaseReader


class ExcelReaderXls(BaseReader):
    def __init__(self):
        super().__init__()

    def read(
        self, file_path: typing.Union[pathlib.Path, typing.IO[bytes]]
    ) -> str:
        workbook = xlrd.open_workbook(file_path)

        sheets_data = []
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)

            rows_data = []
            for row_index in range(sheet.nrows):
                row_data = [
                    str(sheet.cell_value(row_index, i))
                    for i in range(sheet.ncols)
                ]
                rows_data.append(",".join(row_data))

            sheet_content = "\n".join(rows_data)
            sheets_data.append(f"{sheet.name}\n{sheet_content}\n")

        content = "\n".join(sheets_data)
        return content


class ExcelReaderXlsx(BaseReader):
    def __init__(self):
        super().__init__()

    def read(
        self, file_path: typing.Union[pathlib.Path, typing.IO[bytes]]
    ) -> str:
        workbook = openpyxl.load_workbook(file_path)
        sheets_data = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_data = []

            for row in sheet.iter_rows(values_only=True):
                row_data = ",".join(map(str, row))
                rows_data.append(row_data)

            sheet_content = "\n".join(rows_data)
            sheets_data.append(f"{sheet_name}\n{sheet_content}\n")

        content = "\n".join(sheets_data)
        return content
