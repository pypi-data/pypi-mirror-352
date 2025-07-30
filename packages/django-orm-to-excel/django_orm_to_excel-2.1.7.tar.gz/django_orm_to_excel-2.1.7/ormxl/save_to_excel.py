import typing as t

from openpyxl import Workbook
from openpyxl.styles import Font
from django.db.models import Model, QuerySet


def save_to_excel(dest, obj: t.Any, fields: list[str]):
    wb = Workbook()
    ws = wb.active

    if isinstance(obj, QuerySet):
        for col_num, column_title in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_num, value=getattr(obj.model, column_title).field.verbose_name)
            cell.font = Font(bold=True)

        for row_num, row_data in enumerate(obj, 2):
            for col_num, cell_value in enumerate([getattr(row_data, field) for field in fields], 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
    if isinstance(obj, Model):
        for col_num, column_title in enumerate(fields, 1):
            cell = ws.cell(row=col_num, column=1, value=getattr(obj._meta.model, column_title).field.verbose_name)
            cell.font = Font(bold=True)

        for col_num, cell_value in enumerate([getattr(obj, field) for field in fields], 1):
            ws.cell(row=col_num, column=2, value=cell_value)

    wb.save(dest)
