"""
This module provides a toolkit for interacting with Microsoft Graph API to manage
SharePoint sites and files.
It includes functionalities to extract text from files, search files, and list
files and folders in a path, using asynchronous methods.

Classes:
    MSSiteToolKit: A toolkit class for managing SharePoint sites and files using
    Microsoft Graph API.
"""

import io
import openpyxl
from agent_builder.builders.tool_builder import ToolBuilder
from docx import Document
from msgraph import GraphServiceClient
from msgraph.generated.sites.sites_request_builder import SitesRequestBuilder
from PyPDF2 import PdfReader

from recall_space_agents.toolkits.ms_site.schema_mappings import \
    schema_mappings


class MSSiteToolKit:
    """
    A toolkit class for managing SharePoint sites and files using Microsoft Graph API.

    Methods:
        aextract_text_from_file_by_path: Asynchronously extract text from a file 
        given the site name and file path. 
        asearch_and_extract_text: Asynchronously search for a file by name and 
        extract its text content.
        alist_files_and_folders_in_path: Asynchronously list files and folders in 
        a specified path.
        get_tools: Retrieve a list of tools mapped to the methods in the toolkit. 
        Use it to bind tools to agents.
    """

    def __init__(self, credentials):
        """
        Initialize the MSSiteToolKit with Microsoft Graph API client.

        Args:
            credentials: The credentials required to authenticate with 
            the Microsoft Graph API.
        """

        self.required_scopes_as_user = ["Sites.Read.All", "Sites.ReadWrite.All", "Files.Read.All", "Files.ReadWrite.All"]
        self.ms_graph_client = GraphServiceClient(
            credentials=credentials, scopes=self.required_scopes_as_user
        )
        self.schema_mappings = schema_mappings

    async def get_site_id(self, display_name):
        """
        Helper method to get the site ID by it's display name.

        Args:
            display_name (str): The display name of the SharePoint site.

        Returns:
            str or None: The site ID if found; otherwise, None.
        """
        query_params = SitesRequestBuilder.SitesRequestBuilderGetQueryParameters(
            select=["id", "displayName"], search="sites"
        )

        request_config = SitesRequestBuilder.SitesRequestBuilderGetRequestConfiguration(
            query_parameters=query_params
        )
        available_sites = await self.ms_graph_client.sites.get(
            request_configuration=request_config
        )
        for site in available_sites.value:
            if site.display_name == display_name:
                return site.id
        return None

    async def get_drive_id(self, site_id):
        """
        Helper method to get the drive ID by site ID.

        Args:
            site_id (str): The ID of the SharePoint site.

        Returns:
            str or None: The drive ID if found; otherwise, None.
        """
        drive = await self.ms_graph_client.sites.by_site_id(site_id).drives.get()
        return drive.value[0].id if drive.value else None

    async def get_file_id_by_path(self, drive_id, file_path):
        """
        Helper method to get the file ID by path.

        Args:
            drive_id (str): The ID of the drive.
            file_path (str): The path to the file.

        Returns:
            str or None: The file ID if found; otherwise, None.
        """
        file_path = self._remove_document_prefix(file_path)
        folders = file_path.strip("/").split("/")
        file_name = folders.pop()

        folder_id = "root"
        for folder in folders:
            children = (
                await self.ms_graph_client.drives.by_drive_id(drive_id)
                .items.by_drive_item_id(folder_id)
                .children.get()
            )
            folder_id = None
            for child in children.value:
                if child.name == folder:
                    folder_id = child.id
                    break
            if not folder_id:
                return None

        children = (
            await self.ms_graph_client.drives.by_drive_id(drive_id)
            .items.by_drive_item_id(folder_id)
            .children.get()
        )
        file_id = None
        for child in children.value:
            if child.name == file_name:
                file_id = child.id
                break
        return file_id

    async def get_file_content(self, drive_id, file_id):
        """
        Helper method to get the file content by drive ID and file ID.

        Args:
            drive_id (str): The ID of the drive.
            file_id (str): The ID of the file.

        Returns:
            bytes: The binary content of the file.
        """
        file_content = (
            await self.ms_graph_client.drives.by_drive_id(drive_id)
            .items.by_drive_item_id(file_id)
            .content.get()
        )
        return file_content

    async def aextract_text_from_file_by_path(
        self, site_display_name: str, file_path: str
    ):
        """
        Asynchronously extract text from a file given the site name and file path.

        Args:
            site_display_name (str): The display name of the SharePoint site.
            file_path (str): The path to the file within the SharePoint site,
            starting from the root.

        Returns:
            str: The extracted text content of the file, or an error message if
            the operation fails.
        """
        file_path = self._remove_document_prefix(file_path)
        site_id = await self.get_site_id(site_display_name)
        if not site_id:
            return f"Site with display name '{site_display_name}' not found."

        drive_id = await self.get_drive_id(site_id)
        if not drive_id:
            return f"Drive not found for site with ID '{site_id}'."

        file_id = await self.get_file_id_by_path(drive_id, file_path)
        if not file_id:
            return f"File with path '{file_path}' not found."

        file_content = await self.get_file_content(drive_id, file_id)

        # Determine file type and extract text
        if file_path.lower().endswith(".pdf"):
            extracted_text = self.extract_text_from_pdf(file_content)
        elif file_path.lower().endswith(".docx"):
            extracted_text = self.extract_text_from_docx(file_content)
        elif file_path.lower().endswith(".xlsx"):
            extracted_text = self.extract_text_from_xlsx(file_content)
        else:
            return f"Unsupported file type for '{file_path}'."

        return extracted_text

    async def asearch_and_extract_text(self, site_display_name: str, file_name: str):
        """
        Asynchronously search for a file by name and extract its text content.

        Args:
            site_display_name (str): The display name of the SharePoint site.
            file_name (str): The exact name of the file to search for.

        Returns:
            str: The extracted text content of the file, or an error message
            if the operation fails.
        """
        site_id = await self.get_site_id(site_display_name)
        if not site_id:
            return f"Site with display name '{site_display_name}' not found."

        drive_id = await self.get_drive_id(site_id)
        if not drive_id:
            return f"Drive not found for site with ID '{site_id}'."

        search_results = (
            await self.ms_graph_client.drives.by_drive_id(drive_id)
            .search_with_q(q=file_name)
            .get()
        )

        exact_matches = [
            item for item in search_results.value if item.name == file_name
        ]
        if not exact_matches:
            return f"File '{file_name}' not found in site '{site_display_name}'."

        file_item = exact_matches[0]
        file_id = file_item.id

        file_content = await self.get_file_content(drive_id, file_id)

        # Determine file type and extract text
        if file_name.lower().endswith(".pdf"):
            extracted_text = self.extract_text_from_pdf(file_content)
        elif file_name.lower().endswith(".docx"):
            extracted_text = self.extract_text_from_docx(file_content)
        elif file_name.lower().endswith(".xlsx"):
            extracted_text = self.extract_text_from_xlsx(file_content)
        else:
            return f"Unsupported file type for '{file_name}'."

        return extracted_text

    async def alist_files_and_folders_in_path(
        self, site_display_name: str, folder_path: str
    ):
        """
        Asynchronously list files and folders in a specified path.

        Args:
            site_display_name (str): The display name of the SharePoint site.
            folder_path (str): The path to the folder within the SharePoint site,
            starting from the root.

        Returns:
            list or str: A list of item names within the folder, or an error
            message if the operation fails.
        """
        folder_path = self._remove_document_prefix(folder_path)
        site_id = await self.get_site_id(site_display_name)
        if not site_id:
            return f"Site with display name '{site_display_name}' not found."

        drive_id = await self.get_drive_id(site_id)
        if not drive_id:
            return f"Drive not found for site with ID '{site_id}'."

        folders = folder_path.strip("/").split("/")

        folder_id = "root"
        for folder in folders:
            children = (
                await self.ms_graph_client.drives.by_drive_id(drive_id)
                .items.by_drive_item_id(folder_id)
                .children.get()
            )
            folder_id = None
            for child in children.value:
                if child.name == folder:
                    folder_id = child.id
                    break
            if not folder_id:
                return f"Folder '{folder}' not found."

        children = (
            await self.ms_graph_client.drives.by_drive_id(drive_id)
            .items.by_drive_item_id(folder_id)
            .children.get()
        )
        item_names = [child.name for child in children.value]
        return item_names

    def extract_text_from_pdf(self, binary_content):
        """
        Helper method to extract text from a PDF file.

        Args:
            binary_content (bytes): The binary content of the PDF file.

        Returns:
            str: The extracted text content of the PDF.
        """
        with io.BytesIO(binary_content) as f:
            reader = PdfReader(f)
            text = ""
            for page in reader.pages:
                text += page.extract_text()
        return text

    def extract_text_from_docx(self, binary_content):
        """
        Helper method to extract text from a DOCX file.

        Args:
            binary_content (bytes): The binary content of the DOCX file.

        Returns:
            str: The extracted text content of the DOCX.
        """
        with io.BytesIO(binary_content) as f:
            document = Document(f)
            text = "\n".join([para.text for para in document.paragraphs])
        return text

    def extract_text_from_xlsx(self, binary_content):
        """
        Helper method to extract text from an XLSX file.

        Args:
            binary_content (bytes): The binary content of the XLSX file.

        Returns:
            str: The extracted text content of the XLSX.
        """
        with io.BytesIO(binary_content) as f:
            workbook = openpyxl.load_workbook(f, data_only=True)
            text = ""
            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join(
                        [str(cell) if cell is not None else "" for cell in row]
                    )
                    text += f"{row_text}\n"
        return text

    def get_tools(self):
        """
        Retrieve a list of tools mapped to the methods in the toolkit.
        Use it to bind tools to agents.

        Returns:
            list: A list of ToolBuilder objects, each representing a
            method in the toolkit.
        """
        tools = []
        for each_method_key, each_method_value in self.schema_mappings.items():
            tool_builder = ToolBuilder()
            tool_builder.set_name(name=each_method_key)
            tool_builder.set_function(eval(f"self.{each_method_key}"))
            tool_builder.set_coroutine(eval(f"self.{each_method_key}"))
            tool_builder.set_description(description=each_method_value["description"])
            tool_builder.set_schema(schema=each_method_value["input_schema"])
            tool_builder = tool_builder.build()
            tools.append(tool_builder)
        return tools

    def _remove_document_prefix(self, path: str):
        """
        Helper method to normalize the file path by removing prefixes.

        Args:
            path (str): The file or folder path.

        Returns:
            str: The normalized path.
        """
        path = path.lstrip("/")
        path = path.removeprefix("Documents")
        path = path.lstrip("/")
        return f"/{path}"