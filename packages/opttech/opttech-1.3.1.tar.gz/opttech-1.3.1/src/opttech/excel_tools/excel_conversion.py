import re, os
from pathlib import Path
from typing import Iterable, Mapping

import polars as pl
import xlsxwriter
import openpyxl

EXCEL_SHEET_MAX_LEN = 31
DEFAULT_INDEX_SHEET = "RESUMO"
_ILLEGAL = re.compile(r"[:\\/?*\[\]]+")


###############################################################################
# Utilidades
###############################################################################
def _safe_sheet_name(
    raw: str,
    *,
    seen: set[str],
    dup_sep: str = "_",
    trunc_suffix: str = "…",
) -> str:
    """Sanitiza, garante unicidade e ≤31 caracteres."""
    name = _ILLEGAL.sub(" ", raw).strip() or "Sheet"
    if len(name) > EXCEL_SHEET_MAX_LEN:
        keep = EXCEL_SHEET_MAX_LEN - len(trunc_suffix)
        name = name[:keep].rstrip() + trunc_suffix

    base, i = name, 1
    while name in seen:
        extra = f"{dup_sep}{i}"
        name = (base[: EXCEL_SHEET_MAX_LEN - len(extra)] + extra)[:EXCEL_SHEET_MAX_LEN]
        i += 1
    seen.add(name)
    return name


def _load_file(path: Path, *, csv_delimiter: str = ",") -> pl.DataFrame:
    ext = path.suffix.lower()
    if ext == ".csv":
        df = pl.read_csv(
            path,
            separator=csv_delimiter,
            infer_schema_length=0,
            try_parse_dates=False
        )
        # Converte todas as colunas para Utf8 (str)
        return df.with_columns([pl.col(col).cast(pl.Utf8) for col in df.columns])

    
    if ext == ".parquet":
        return pl.read_parquet(path)
    
    if ext in {".xls", ".xlsx"}:
        # Excel → openpyxl → polars
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        dfs = []
        for ws in wb.worksheets:
            data = [list(row) for row in ws.iter_rows(values_only=True)]
            if not data:
                continue
            header, *rows = data
            colnames = [
                str(c).strip() if c is not None else f"col_{i}"
                for i, c in enumerate(header)
            ]
            rows = [[str(cell) if cell is not None else "" for cell in row] for row in rows]
            df = pl.DataFrame(rows, schema=[(name, pl.Utf8) for name in colnames])
            dfs.append(df.with_columns(pl.lit(ws.title).alias("__sheet__")))
        return pl.concat(dfs)
    raise ValueError(f"Extensão não suportada: {path}")


###############################################################################
# API de alto nível
###############################################################################
class WorkbookSpec:
    """Parâmetros declarativos para geração de um workbook."""

    def __init__(
        self,
        output_path: str | Path,
        input_files: Iterable[str | Path],
        *,
        index_sheet_name: str = DEFAULT_INDEX_SHEET,
        trunc_suffix: str = "…",
        csv_delimiter: str = "|",

    ) -> None:
        self.output_path = Path(output_path)
        self.index_sheet_name = index_sheet_name
        self.trunc_suffix = trunc_suffix
        self.csv_delimiter = csv_delimiter

        # Expande globs
        self.input_files: list[Path] = []
        for p in input_files:
            path = Path(p)
            if "*" in path.name:
                self.input_files.extend(sorted(path.parent.glob(path.name)))
            else:
                self.input_files.append(path)


def write_xlsx(spec: WorkbookSpec) -> Path:
    seen: set[str] = set()
    index_rows: list[Mapping[str, object]] = []
    ws_index = None  # Referência opcional à aba de índice

    with xlsxwriter.Workbook(spec.output_path) as wb:
        # 1. Se mais de um arquivo, cria primeiro a aba de índice (ainda vazia)
        if len(spec.input_files) > 1:
            index_name = _safe_sheet_name(
                spec.index_sheet_name, seen=seen, trunc_suffix=spec.trunc_suffix
            )
            ws_index = wb.add_worksheet(index_name)

        # 2. Agora processa os arquivos
        for path in spec.input_files:
            df = _load_file(path, csv_delimiter=spec.csv_delimiter)
            sheet = _safe_sheet_name(path.stem, seen=seen, trunc_suffix=spec.trunc_suffix)

            ws = wb.add_worksheet(sheet)

            for col, name in enumerate(df.columns):
                ws.write(0, col, name)
                ws.write_column(1, col, df[name].to_list())

            index_rows.append({
                "Arquivo": path.name,
                "Aba": sheet,
                "Linhas": df.height,
                "Colunas": df.width,
            })

        # 3. Preenche a aba de índice (se criada)
        if ws_index:
            headers = ["Arquivo", "Aba", "Linhas", "Colunas"]
            for col, h in enumerate(headers):
                ws_index.write(0, col, h)
            for row_idx, row in enumerate(index_rows, start=1):
                for col_idx, h in enumerate(headers):
                    ws_index.write(row_idx, col_idx, row[h])

    return spec.output_path


def read_xlsx(path: str | Path) -> dict[str, pl.DataFrame]:
    """
    Lê TODAS as abas de um `.xlsx` e devolve em dict[str, pl.DataFrame].
    Implementado via `openpyxl` → listas → `pl.DataFrame`.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: dict[str, pl.DataFrame] = {}

    for ws in wb.worksheets:
        data = [list(row) for row in ws.iter_rows(values_only=True)]
        if not data or all(cell is None for cell in data[0]):
            continue

        header, *rows = data

        # Sanitize cabeçalho
        colnames = [
            str(c).strip() if c is not None else f"col_{i}"
            for i, c in enumerate(header)
        ]

        # Converte todas as células para str (ou "")
        str_rows = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in rows
        ]

        # Cria DataFrame com tipo str
        try:
            df = pl.DataFrame(str_rows, schema=[(name, pl.Utf8) for name in colnames])
        except Exception as e:
            raise ValueError(f"Erro ao processar aba '{ws.title}': {e}")

        out[ws.title] = df

    return out


###############################################################################
# Exemplo de uso
###############################################################################
if __name__ == "__main__" :

    folder_path = "/home/eduardo/Packages/opttech/tests/data/ExcelDataTeste"
    abas = read_xlsx(os.path.join(folder_path, "Teste.xlsx"))

    spec = WorkbookSpec(
        output_path= f"{folder_path}/relatorio_polars.xlsx",
        input_files= [
            f"{folder_path}/*.csv", # Todos os CSVs da pasta
            f"{folder_path}/*.parquet" # Arquivo específico
        ]
    )
    destino = write_xlsx(spec)
    print("Arquivo gerado em:", Path(destino).resolve())
