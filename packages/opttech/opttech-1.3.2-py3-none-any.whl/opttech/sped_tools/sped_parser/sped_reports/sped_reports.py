# External Models
import os, csv, json
from copy import deepcopy
import openpyxl
from openpyxl.styles import PatternFill, Font
from copy import copy
import pickle
import shutil
from datetime import datetime

from ....tools import detect_encoding, convert_csv2parquet

from ....parameters.sped.constants import *
from ....parameters.constants import *
from ...utils import generate_hash_and_signdate, load_sped_dictionary
from .dt_proc import get_dt_trans

# from .sped_parser.parser_tools import io_tools



def get_storage_filename( file_path : str ) -> str:
    
    if '/FILES_RECIEVED/' in file_path:
        _file_path = file_path.split('/')
        request_id = None
        for v_id, v in enumerate(_file_path):
            if v == 'FILES_RECIEVED':
                request_id = _file_path[v_id-1]
                break
        
        # _file_path = file_path.split(request_id)
        _file_path = file_path.split('/FILES_RECIEVED/')
        _file_path = _file_path[-1]
        _file_path = _file_path.replace(f'{request_id}_', '')
        _file_path = _file_path.replace(request_id, '')
        for part in request_id.split('_'):
            _file_path = _file_path.replace(part, '')
        
        return _file_path
    else:
        return file_path



def load_compl_tables():
    filenames = os.listdir(TABLES_FOLDER)

    compl_tables = {}
    for file_name in filenames:
        file_path = os.path.join(TABLES_FOLDER, file_name)

        f_in = open(file=file_path, mode='r', newline='', encoding=ENCODING)
        reader = csv.DictReader(f_in, delimiter=DELIMITER, quotechar=QUOTECHAR, escapechar='\\')

        fieldnames = list( reader.fieldnames )
        field_cod = fieldnames[0]
        field_desc = fieldnames[1]

        for row_id, row in enumerate(reader):
            if row_id == 0:
                compl_tables[field_cod] = {
                    'field_desc' : field_desc,
                    row[field_cod] : row[field_desc]
                }
            else:
                compl_tables[field_cod][row[field_cod]] = row[field_desc]

        f_in.close()

    return compl_tables



def build_row(
        line : str,
        delimiter : str,
        delimiter_start : bool,
        file : object,
        row_id : int,
        file_info : dict,
        sped_dict : dict,
        logs : list = None
        ) -> list:

    row = None
    if delimiter_start is True:
        
        if len(line.split(delimiter)) == 3:
            line = line.rstrip('\r\n')

        if '\n' not in line:
            str_part = line
            str_parts = [str_part]
            while '\n' not in str_part:
                str_part = next(file, None)
                if str_part is None:
                    return []
                str_parts.append(str_part)
            line = ''.join(str_parts)

        # Treat line
        _line = line.rstrip('\r\n')
        if _line and _line[-1] != delimiter:
            reg = line[1:5]
            if reg not in sped_dict:
                # return False
                raise ValueError(f'ERROR: [sped_parser] - Unpredicted build row interaction (1). Suposto registro "{reg}" não identificado no manual.')

            if 'IND_FIM_RTF' not in sped_dict[reg]['campos']:
                raise ValueError('ERROR: [sped_parser] - Unpredicted build row interaction (2).')
            
            # BUILD FIRST PART
            row = line.split(delimiter)
                
            # GET RTF FILE PARTS
            rtf_part = row[-1]
            end_constant = f"{delimiter}{reg}FIM{delimiter}"
            rtf_parts = [rtf_part]
            while not rtf_part.endswith(end_constant):
                rtf_part  = next(file, None)
                rtf_parts.append(rtf_part)
                if rtf_part is None:
                    return []
                rtf_part = rtf_part.rstrip('\r\n')
            
            # COMBINE RTF PARTS
            rtf_parts[-1] = rtf_parts[-1].rstrip('\r\n')[:-9]
            row[-1] = ''.join(rtf_parts)
            rtf_parts.clear()
            del rtf_parts

            # END OF RTF FILE
            row.append(f"{reg}FIM")
            row.append('')

            # line = line.rstrip('\r\n').replace('\x00', '')
            # row = line.split(delimiter)

        else:
            # build row
            line = line.rstrip('\r\n').replace('\x00', '')
            row = line.split(delimiter)
    else:

        # build row
        line = line.rstrip('\r\n').replace('\x00', '')
        row = f'{delimiter}{line}{delimiter}'.split(delimiter)

    if len(row) <= 1:
        if logs is not None:
            log_message = f'Arquivo fora do padrão sped, evidência na linha {row_id}.'
            # log.write_log( 'build_row', log_message, 'file error' )

            log_row = { 'function' : 'build_row', 'message' : log_message }
            logs.append(log_row)
        return []
    
    # Check row
    if row[0] != '' or row[1] not in sped_dict or row[-1] != '':
        if logs is not None:
            if row[1] in sped_dict:
                log_message = f'Arquivo fora do padrão sped, evidência na linha {row_id}.'
                # log.write_log( 'build_row', log_message, 'file error' )                

                log_row = { 'function' : 'build_row', 'message' : log_message }
                logs.append(log_row)
            else:
                log_message = f'Arquivo com registro {row[1]} não identificado no manual sped correspondente, evidência na linha {row_id}.'
                # log.write_log( 'build_row', log_message, 'file error' )

                log_row = { 'function' : 'build_row', 'message' : log_message }
                logs.append(log_row)
                
            file_info['Processado'] = False

        return []
    
    return row


def check_row(
        line : str,
        delimiter : str,
        delimiter_start : bool,
        file : object,
        sped_dict : dict,
    ) -> bool:  

    row = None
    if delimiter_start is True:
        # Treat line
        _line = line.rstrip('\r\n')
        if _line and _line[-1] != delimiter:
            reg = line[1:5]
            if reg not in sped_dict:
                raise ValueError('ERROR: [sped_parser] - Unpredicted build row interaction (1).')

            if 'IND_FIM_RTF' not in sped_dict[reg]['campos']:
                raise ValueError('ERROR: [sped_parser] - Unpredicted build row interaction (2).')
            
            # BUILD FIRST PART
            row = line.split(delimiter)
                
            # GET RTF FILE PARTS
            rtf_part = row[-1]
            end_constant = f"{delimiter}{reg}FIM{delimiter}"
            rtf_parts = [rtf_part]
            while not rtf_part.endswith(end_constant):
                rtf_part  = next(file, None)
                rtf_parts.append(rtf_part)
                if rtf_part is None:
                    return False
                rtf_part = rtf_part.rstrip('\r\n')
            
            # COMBINE RTF PARTS
            rtf_parts[-1] = rtf_parts[-1].rstrip('\r\n')[:-9]
            row[-1] = ''.join(rtf_parts)
            rtf_parts.clear()
            del rtf_parts

            # END OF RTF FILE
            row.append(f"{reg}FIM")
            row.append('')

            # line = line.rstrip('\r\n').replace('\x00', '')
            # row = line.split(delimiter)

        else:    
            # build row
            line = line.rstrip('\r\n').replace('\x00', '')
            row = line.split(delimiter)
    else:

        # build row
        line = line.rstrip('\r\n').replace('\x00', '')
        row = f'{delimiter}{line}{delimiter}'.split(delimiter)

    if len(row) <= 1:
        return False
    
    # Check row
    if row[0] != '' or row[1] not in sped_dict or row[-1] != '':
        return False
    
    return True



def build_row_dict(
        reg : str,
        row : list,
        sped_dict : dict,
        prefix : str,
        row_id : int,
        compl_tables : dict = {},
        summary_by_field : dict = {},
        logs : list = None
    ) -> tuple[dict,bool]:
    
    # TODO: Optimize the construction of this dictionary (row_dict = {})
    row_dict = {}
    row_error = False

    # if (prefix != 'ECF' or reg != 'Y720') and reg[0] != '9':

    if sped_dict[reg]['num_campos'] != len(row) - 2:
        # if logs is not None:
        #     log_message = f"Esperava-se {sped_dict[reg]['num_campos']} campos no registro {reg} na linha {row_id} do txt e a ferramenta identificou {len(row) - 2}."
        #     log_row = { 'function' : 'build_row_dict', 'message' : log_message, 'error_type' : 'layout error' }
        #     logs.append(log_row)
        
        if sped_dict[reg]['num_campos'] < len(row) - 2:
            row = row[:sped_dict[reg]['num_campos']+2]
            row[-1] = ''

        else:
            num_add = sped_dict[reg]['num_campos'] - (len(row) - 2)
            for i in range(num_add):
                row.append('')
            
            # row_error = True

    # CHECK LENGTH
    if sped_dict[reg]['num_campos'] == len(row) - 2:
        for field in sped_dict[reg]['campos']:
            field_name = f'{reg}{prefix}_{field}'
            value = row[sped_dict[reg]['campos'][field]['ordem']]

            row_dict[field_name] = value

            # DESCR FIELD
            if field in compl_tables:
                if value != '':
                    _field = compl_tables[field]['field_desc']
                    field_name = f'{reg}{prefix}_{_field}'
                    if value in compl_tables[field]:
                        row_dict[field_name] = compl_tables[field][value]


    return row_dict, row_error



def build_simple_row_dict( 
        reg : str, 
        row : list, 
        sped_dict : dict, 
        prefix : str, 
        row_id : int = -2, 
        delimiter_start : bool = True, 
        log = None 
        ) -> dict:    
    
    row_dict = {}

    # CHECK LENGTH
    if sped_dict[reg]['num_campos'] == len(row) - 2 * delimiter_start:
        for field in sped_dict[reg]['campos']:
            field_name = f'{reg}{prefix}_{field}'
            value = row[sped_dict[reg]['campos'][field]['ordem']]

            row_dict[field_name] = value

    else:
        if log is not None:
            log_message = f'Erro na linha: {row_id+1}.'
            log.write_log( 'build_row_dict', log_message, 'layout error' )

    return row_dict


def build_pointers(
        sped : str,
        version : str,
        cnpj : str,
        competencia : str,
        retificador_original : str,
        sped_dict : dict,
        output_path : str,
        report_fieldnames : dict,
        hash_file : str,
        file_name : str
        ) -> bool:

    file_pointers = {}
    for reg in REPORT_REG[sped]:
        if reg not in sped_dict:
            continue
        
        file_path = None
        if sped != 'ECD':
            report_name = f"REPORT-CNPJ_{cnpj}-SPED_{sped}-VERSION_{version}-COMPETENCIA_{competencia.replace('-','')}-REG_{reg}-RETIFICADOR_{retificador_original.upper()}-HASH_{hash_file}.csv"
            file_path = os.path.join( output_path, report_name )
        else:
            leiaute = SPED_BY_VERSION[sped][version]['version']
            file_path = os.path.join( output_path, f"REPORT-CNPJ_{cnpj}-SPED_{sped}-VERSION_{leiaute}-COMPETENCIA_{competencia.replace('-','')}-REG_{reg}-RETIFICADOR_{retificador_original.upper()}-HASH_{hash_file}.csv" )


        f_in = open( file_path, mode='w', newline='', encoding=ENCODING, errors='replace' )
        fieldnames = {
            'CompetenciaArquivo' : 'CompetenciaArquivo',
            **report_fieldnames[reg], 
            'RowId' : 'RowId',
            'HashFile' : 'HashFile', 
            'FileName' : 'FileName'
            }
        writer = csv.DictWriter( f_in, fieldnames=fieldnames, delimiter=DELIMITER, quotechar=QUOTECHAR, quoting=csv.QUOTE_MINIMAL, escapechar='\\' )
        writer.writeheader()
        file_pointers[reg] = {
            'hash_file' : hash_file,
            'file_path' : file_path,
            'file_name' : file_name,
            'fieldnames' : fieldnames,
            'report_level' : REPORT_REG[sped][reg],
            'pointer' : f_in,
            'writer' : writer,
            'num_regs' : 0,
            'num_rows' : 0,
            # 'count_by_field' : {field : 0 for field in fieldnames}
        }

    return file_pointers


def build_reg_pointers(
        sped : str,
        version : str,
        cnpj : str,
        competencia : str,
        retificador_original : str,
        prefix : str,
        sped_dict : dict,
        output_path : str,
        ret_report_fieldnames : dict,
        hash_file : str,
        file_name : str
        ) -> bool:

    # count_path = os.path.join( output_path, 'COUNT_FIELD' )
    # if os.path.exists(count_path) == False:
    #     os.mkdir(count_path)

    reg_file_pointers = {}
    for reg in sped_dict:

        file_path = None
        if sped != 'ECD':
            file_path = os.path.join( output_path, f"REG-CNPJ_{cnpj}-SPED_{sped}-VERSION_{version}-COMPETENCIA_{competencia.replace('-','')}-REG_{reg}-RETIFICADOR_{retificador_original.upper()}-HASH_{hash_file}.csv" )
            # count_file_path = os.path.join( count_path, f"REG-CNPJ_{cnpj}-SPED_{sped}-VERSION_{version}-COMPETENCIA_{competencia.replace('-','')}-REG_{reg}-RETIFICADOR_{retificador_original.upper()}-HASH_{hash_file}.csv" )
        else:
            leiaute = SPED_BY_VERSION[sped][version]['version']
            file_path = os.path.join( output_path, f"REG-CNPJ_{cnpj}-SPED_{sped}-VERSION_{leiaute}-COMPETENCIA_{competencia.replace('-','')}-REG_{reg}-RETIFICADOR_{retificador_original.upper()}-HASH_{hash_file}.csv" )
            # count_file_path = os.path.join( count_path, f"REG-CNPJ_{cnpj}-SPED_{sped}-VERSION_{leiaute}-COMPETENCIA_{competencia.replace('-','')}-REG_{reg}-RETIFICADOR_{retificador_original.upper()}-HASH_{hash_file}.csv" )


        f_in = open( file_path, mode='w', newline='', encoding=ENCODING, errors='replace' )
        # f_in_count = open( count_file_path, mode='w', encoding='UTF-8', newline='', errors='replace' )

        fieldnames = {
            'CompetenciaArquivo' : 'CompetenciaArquivo',
            'CNPJ' : 'CNPJ',
            **{ field : field for field in ret_report_fieldnames[reg]}, 
            'HashFile' : 'HashFile', 
            'FileName' : 'FileName'
            }
        writer = csv.DictWriter( 
            f_in, 
            fieldnames=fieldnames, 
            delimiter=DELIMITER, 
            quotechar=QUOTECHAR, 
            quoting=csv.QUOTE_MINIMAL, 
            escapechar='\\' 
            )
        writer.writeheader()
        # count_writer = csv.DictWriter( f_in_count, delimiter='|', fieldnames=fieldnames, quotechar=QUOTECHAR, quoting=csv.QUOTE_MINIMAL, escapechar='\\' )
        # count_writer.writeheader()

        reg_file_pointers[reg] = {
            'hash_file' : hash_file,
            'file_path' : file_path,
            'file_name' : file_name,
            'fieldnames' : fieldnames,
            'pointer' : f_in,
            # 'count_pointer' : f_in_count,
            'writer' : writer,
            # 'count_writer' : count_writer,
            'num_rows' : 0,
            # 'data' : []
        }

    return reg_file_pointers




def build_list_iterative(son_by_father : dict) -> list:

    # LEVEL 1 ------------ START
    for reg_son_n1 in son_by_father:
        _son_by_father_n1 = son_by_father[reg_son_n1]

        # LEVEL 2 ------------ START
        new_list_l1 = []
        new_list_l1_positions = []
        for reg_son_n2 in _son_by_father_n1['sons']:
            _son_by_father_n2 = _son_by_father_n1['sons'][reg_son_n2]

            # LEVEL 3 ------------ START
            new_list_l2 = []
            new_list_l2_positions = []
            for reg_son_n3 in _son_by_father_n2['sons']:
                _son_by_father_n3 = _son_by_father_n2['sons'][reg_son_n3]

                # LEVEL 4 ------------ START
                new_list_l3 = []
                new_list_l3_positions = []
                for reg_son_n4 in _son_by_father_n3['sons']:
                    _son_by_father_n4 = _son_by_father_n3['sons'][reg_son_n4]

                    # list (3)
                    for l3_id, row_l3 in enumerate(_son_by_father_n3['rows']):
                        _insert_n4 = False
                        for l4_id, row_l4 in enumerate(_son_by_father_n4['rows']):
                            if l3_id == _son_by_father_n4['positions'][l4_id]:
                                _insert_n4 = True
                                new_list_l3.append( { **row_l3, **row_l4 } )
                                new_list_l3_positions.append(_son_by_father_n3['positions'][l3_id])
                        
                        if not _insert_n4:
                            new_list_l3.append( copy(row_l3) )
                            new_list_l3_positions.append(_son_by_father_n3['positions'][l3_id])

                    _son_by_father_n4.clear()
                    _son_by_father_n4 = None

                if new_list_l3 != []:
                    _son_by_father_n3['rows'] = new_list_l3
                    _son_by_father_n3['positions'] = new_list_l3_positions
                    _son_by_father_n3['size'] = len(new_list_l3)

                # LEVEL 4 ------------ END


                # list (2)
                for l2_id, row_l2 in enumerate(_son_by_father_n2['rows']):
                    _insert_n3 = False
                    for l3_id, row_l3 in enumerate(_son_by_father_n3['rows']):
                        if l2_id == _son_by_father_n3['positions'][l3_id]:
                            _insert_n3 = True
                            new_list_l2.append( { **row_l2, **row_l3 } )
                            new_list_l2_positions.append(_son_by_father_n2['positions'][l2_id])
                    
                    if not _insert_n3:
                        new_list_l2.append( copy(row_l2) )
                        new_list_l2_positions.append(_son_by_father_n2['positions'][l2_id])

                _son_by_father_n3.clear()
                _son_by_father_n3 = None

            if new_list_l2 != []:
                _son_by_father_n2['rows'] = new_list_l2
                _son_by_father_n2['positions'] = new_list_l2_positions
                _son_by_father_n2['size'] = len(new_list_l2)

            # LEVEL 3 ------------ END


            # list (1)
            for l1_id, row_l1 in enumerate(_son_by_father_n1['rows']):
                _insert_n2 = False
                for l2_id, row_l2 in enumerate(_son_by_father_n2['rows']):
                    if l1_id == _son_by_father_n2['positions'][l2_id]:
                        _insert_n2 = True
                        new_list_l1.append( { **row_l1, **row_l2 } )
                        new_list_l1_positions.append(_son_by_father_n1['positions'][l1_id])
                
                if not _insert_n2:
                    new_list_l1.append( copy(row_l1) )
                    new_list_l1_positions.append(_son_by_father_n1['positions'][l1_id])

            # # list (1)
            # for l2_id, row_l2 in enumerate( _son_by_father_n2['rows'] ):
            #     _position = _son_by_father_n2['positions'][l2_id]
            #     new_list_l1.append({ **_son_by_father_n1['rows'][_position], **row_l2 })
            #     new_list_l1_positions.append(_son_by_father_n1['positions'][_position])

            _son_by_father_n2.clear()
            _son_by_father_n2 = None

        if new_list_l1 != []:
            _son_by_father_n1['rows'] = new_list_l1
            _son_by_father_n1['positions'] = new_list_l1_positions
            _son_by_father_n1['size'] = len(new_list_l1)

        # LEVEL 2 ------------ END
    
    
    # Get the number of rows for the current level
    num_rows = 0
    for reg_son in son_by_father:
        if num_rows < son_by_father[reg_son]['size']:
            num_rows = son_by_father[reg_son]['size']

    # Join regs
    list_to_write = []
    for row_id in range(num_rows):
        _dict_to_write = {}
        for reg_son in son_by_father:
            if row_id < son_by_father[reg_son]['size']:
                _dict_to_write = {
                    **_dict_to_write,
                    **son_by_father[reg_son]['rows'][row_id]
                }

        list_to_write.append( _dict_to_write )

    return list_to_write



def write_row(
        reg : str, 
        file_pointers : dict, 
        report_reg_compl : dict, 
        son_by_father : dict, 
        father_by_reg : dict, 
        allow_father_by_reg : dict, 
        compl_info : dict, 
        summary_by_report : dict,
        file_info : dict
        ):
    
    if allow_father_by_reg[reg] is False:
        return False

    if reg not in file_pointers:
        return False

    list_to_write = []

    # Consolidate children
    if reg in son_by_father['sons'] and son_by_father['sons'][reg]['sons'] != {}:

        list_to_write = build_list_iterative( son_by_father['sons'][reg]['sons'] )
        for row_id, row in enumerate(list_to_write):
            list_to_write[row_id] = {
                **father_by_reg[reg],
                **list_to_write[row_id]
            }

    if list_to_write == []:
        list_to_write = [father_by_reg[reg]]

    # Complement REG INFO (Tabelas)
    list_lenth = len(list_to_write)
    for compl_reg in report_reg_compl:
        if reg not in report_reg_compl[compl_reg]:
            continue

        index = 0
        while index < list_lenth:
            key_value = get_compl_key( reg, report_reg_compl[compl_reg], list_to_write[index] )
            if key_value in compl_info[compl_reg]:

                if 'filter' in report_reg_compl[compl_reg]:
                    list_to_write[index] = {
                        **list_to_write[index],
                        **{ field : compl_info[compl_reg][key_value][field] for field in compl_info[compl_reg][key_value] if field not in report_reg_compl[compl_reg]['filter'] }
                    }
                else:
                    list_to_write[index] = {
                        **list_to_write[index],
                        **compl_info[compl_reg][key_value]
                    }

            else:
                # TODO: Incluir tratamento de erro
                pass
            index += 1


    # Writer report row
    for _row in list_to_write:
        file_pointers[reg]['num_rows'] += 1
        row = {
            'CompetenciaArquivo' : file_info['Competencia'],
            **_row,
            'RowId' : f"ID_{file_pointers[reg]['num_rows']}",
            'HashFile' : file_pointers[reg]['hash_file'], 
            'FileName' : file_pointers[reg]['file_name']
            }

        file_pointers[reg]['writer'].writerow( row )

    file_pointers[reg]['num_regs'] += 1
    

    # Clear
    if reg in son_by_father['sons']:
        son_by_father['sons'][reg].clear()
        son_by_father['sons'][reg] = {
            'rows' : [],
            'size' : 1,
            'positions' : [],
            'sons' : {}
        }
        # gc.collect()
    
    return True


def get_compl_key( reg : str, report_reg_compl : list, reg_dict : dict ) -> str:
    key_value = ''
    for field in report_reg_compl[reg]:
        if field not in reg_dict:
            return False
        
        key_value += f'|{reg_dict[field]}'
    return key_value


def get_report_fieldnames(
        sped : str, 
        prefix : str,
        sped_dict : dict, 
        chield_level : int, 
        report_fieldnames : dict, 
        report_reg_compl : dict,
        compl_tables : dict,
        report_fieldnames_types : dict = None,
    ) -> dict:

    # TODO: Memoizar report_fieldnames

    fieldname_dict = {'' : {}}

    fieldname_dict_type = None
    if report_fieldnames_types is not None:
        fieldname_dict_type = {'' : {}}
    
    for reg in sped_dict:
        reg_father = sped_dict[reg]['pai']

        # field_dict = { f'{reg}{prefix}_{field}' : field for field in sped_dict[reg]['campos'] }

        field_dict = {}
        for field in sped_dict[reg]['campos']:
            field_name = f'{reg}{prefix}_{field}'
            field_dict[field_name] = field

            if field in compl_tables:
                _field = compl_tables[field]['field_desc']
                field_name = f'{reg}{prefix}_{_field}'
                field_dict[field_name] = _field

        field_dict_type = {}
        for field in sped_dict[reg]['campos']:
            field_name = f'{reg}{prefix}_{field}'

            if "DT_" in field and sped_dict[reg]['campos'][field]['tam'] == 8:
                field_dict_type[field_name] = 'DATE'

            elif 'CNPJ' in field or 'CPF' in field:
                field_dict_type[field_name] = 'STRING'

            else:
                if "DT_" in field and sped_dict[reg]['campos'][field]['tam'] != 8:
                    ...
                field_dict_type[field_name] = 'STRING' if sped_dict[reg]['campos'][field]['tipo'] == 'text' else sped_dict[reg]['campos'][field]['tipo'].upper()

            if field in compl_tables:
                _field = compl_tables[field]['field_desc']
                field_name = f'{reg}{prefix}_{_field}'
                field_dict_type[field_name] = 'STRING'

        fieldname_dict[reg] = {
            **fieldname_dict[reg_father], 
            **field_dict
            }
        
        if report_fieldnames_types is not None:
            fieldname_dict_type[reg] = {
                **fieldname_dict_type[reg_father], 
                **field_dict_type
            }

        if reg in REPORT_REG[sped]:
            bloco = reg[0]
            report_fieldnames[reg] = deepcopy(fieldname_dict[reg])
            if report_fieldnames_types is not None:
                report_fieldnames_types[reg] = deepcopy(fieldname_dict_type[reg])

        # Get chield
        cl = 1
        current_reg = reg
        while cl <= chield_level:
            # Get next father
            reg_father = sped_dict[current_reg]['pai']
            if reg_father == '':
                break

            if reg_father in REPORT_REG[sped] and cl <= REPORT_REG[sped][reg_father]:
                report_fieldnames[reg_father] = {
                    **report_fieldnames[reg_father], 
                    **fieldname_dict[reg]
                    }
                if report_fieldnames_types is not None:
                    report_fieldnames_types[reg_father] = {
                        **report_fieldnames_types[reg_father], 
                        **fieldname_dict_type[reg]
                        }
                
            current_reg = reg_father
            cl += 1


    # Add compl fields
    for reg in REPORT_REG[sped]:
        if reg not in sped_dict:
            continue

        if reg in report_reg_compl[sped]:
            continue

        bloco = reg[0]
        # Compl
        for reg_compl in report_reg_compl[sped]:
            if reg_compl not in sped_dict:
                continue
            
            if reg not in report_reg_compl[sped][reg_compl]:
                report_reg_compl[sped][reg_compl][reg] = []
            
            for field_ in report_reg_compl[sped][reg_compl]['fields']:
                if field_[0] != '*':
                    field = field_

                    # find field in BLOCO
                    for _field in report_fieldnames[reg]:
                        if bloco == _field[0] and field == report_fieldnames[reg][_field]:
                            report_reg_compl[sped][reg_compl][reg].append(_field)
                            break

                    
                    # for _field in report_fieldnames_types[reg]:
                    #     if bloco == _field[0] and field == report_fieldnames_types[reg][_field]:
                    #         report_reg_compl[sped][reg_compl][reg].append(_field)
                    #         break

                else:
                    field = field_[1:]

                    # find field from begining
                    for _field in report_fieldnames[reg]:
                        if field == report_fieldnames[reg][_field]:
                            report_reg_compl[sped][reg_compl][reg].append(_field)
                            break
                    
                    # for _field in report_fieldnames_types[reg]:
                    #     if field == report_fieldnames_types[reg][_field]:
                    #         report_reg_compl[sped][reg_compl][reg].append(_field)
                    #         break

    # Clean reg compl
    for reg_compl in report_reg_compl[sped]:
        if reg_compl not in sped_dict:
            continue

        num_fields = len(report_reg_compl[sped][reg_compl]['fields'])
        for reg in REPORT_REG[sped]:
            if reg not in sped_dict:
                continue
            if reg not in report_reg_compl[sped][reg_compl]:
                continue
            if num_fields != len(report_reg_compl[sped][reg_compl][reg]):
                del report_reg_compl[sped][reg_compl][reg]

    # Complement fieldnames
    for reg_compl in report_reg_compl[sped]:
        if reg_compl not in sped_dict:
            continue

        for reg in report_reg_compl[sped][reg_compl]:
            if reg == reg_compl or reg == 'fields' or reg == 'filter':
                continue

            if 'filter' in report_reg_compl[sped][reg_compl]:
                report_fieldnames[reg] = {
                    **{ field : fieldname_dict[reg_compl][field] for field in fieldname_dict[reg_compl] if field not in report_reg_compl[sped][reg_compl]['filter'] },
                    **report_fieldnames[reg]
                    }
                
                if report_fieldnames_types is not None:
                    report_fieldnames_types[reg] = {
                        **{ field : fieldname_dict_type[reg_compl][field] for field in fieldname_dict_type[reg_compl] if field not in report_reg_compl[sped][reg_compl]['filter'] },
                        **report_fieldnames_types[reg]
                        }

            else:
                report_fieldnames[reg] = {
                    **fieldname_dict[reg_compl],
                    **report_fieldnames[reg]
                    }

                if report_fieldnames_types is not None:
                    report_fieldnames_types[reg] = {
                        **fieldname_dict_type[reg_compl],
                        **report_fieldnames_types[reg]
                        }

    # # Sort Fieldnames
    # for reg in report_fieldnames:
    #     report_fieldnames[reg] = dict(sorted(report_fieldnames[reg].items(), key=lambda item : item[0]))

    if report_fieldnames_types is not None:
        return report_fieldnames, report_fieldnames_types
    
    return report_fieldnames



def get_reg_report_fieldnames( 
        prefix : str,
        sped_dict : dict, 
        compl_tables : dict
        ) -> dict:

    # TODO: Memoizar report_fieldnames

    fieldname_dict = {}
    for reg in sped_dict:

        field_dict = {}
        for field in sped_dict[reg]['campos']:
            field_name = f'{reg}{prefix}_{field}'
            field_dict[field_name] = field

            if field in compl_tables:
                _field = compl_tables[field]['field_desc']
                field_name = f'{reg}{prefix}_{_field}'
                field_dict[field_name] = _field

        fieldname_dict[reg] = field_dict

    return fieldname_dict



def get_sped_version( row : list, sped : str ):
    if sped == 'EFD_CONTR':
        return row[2]
    
    elif sped == 'EFD_FISCAL':
        return row[2]
    
    elif sped == 'ECF':
        return row[3]
    
    elif sped == 'ECD':
        return row[4][-4:]
    
    elif sped == 'MANAD':
        return row[14]

    return None



# DEV
def get_sped_info( 
        row : list, 
        sped_by_version : dict, 
        logs : list = None, 
        file_path : str = None, 
        detected_encoding : str = None, 
        delimiter : str = None,
    ) -> bool:

    sped = classify_sped( row )
    version = get_sped_version( row, sped )
    if version is None:
        return False, False, False, False
    
    if version not in sped_by_version[sped]:
        if logs is not None:
            log_message = f'Versão leiaute {version} inexistente. Arquivo de {row[4]}.'
            log_row = {'function' : 'get_sped_info', 'message' : log_message}
            logs.append(log_row)
        return False, False, False, False

    prefixo = sped_by_version[sped][version]['prefixo']
    delimiter_start = sped_by_version[sped][version]['delimiter_start']

    return version, sped, prefixo, delimiter_start



def get_parents( reg : str, sped_dict : dict, chield_level : int, REPORT_REG : dict,  ) -> list:
    related_parents = []

    level = 1
    _reg = reg

    _reg_path = []
    while level <= chield_level:
        reg_father = sped_dict[_reg]['pai']

        if reg_father == '':
            break

        _reg_path.append(reg_father)

        if reg_father in REPORT_REG and level <= REPORT_REG[reg_father]:
            _list = deepcopy(_reg_path)
            _list.reverse()
            related_parents.append(_list)

        _reg = reg_father
        level += 1

    return related_parents




def classify_sped( row : list ) -> str:
    # Não tem pipe no início
    # Os campos 13 e 14 estão como data

    def is_valid_date( date_str ):
        if len( date_str ) != 8:
            return False
        try:
            day, month, year = int(date_str[:2]), int(date_str[2:4]), int(date_str[4:])
            return 1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= 2100
        except:
            return False
    
    if type(row) != list:
        return None
    
    if len(row) < 2:
        return None


    # Verifica se a string começa com "|0000|"
    if row[1] == "0000":
        if "LECD" in row[2]:
            # TODO: Ajustar
            return "ECD"
        
        elif "LECF" in row[2]:
            # TODO: Ajustar
            return "ECF"
        
        elif is_valid_date(row[6]) and is_valid_date(row[7]):
            return "EFD_CONTR"
        
        elif is_valid_date(row[4]) and is_valid_date(row[5]):
            return "EFD_FISCAL"
        else:
            return None

    elif row[0] == '0000':
        if is_valid_date(row[12]) and is_valid_date(row[13]):
            return 'MANAD'

    return None


def get_cnpj( row : list, sped : str, report_by_cnpj : bool ) -> str:

    if report_by_cnpj == True:
        if sped == 'MANAD':
            return row[3]
        elif sped == 'EFD_FISCAL':
            return None
        elif sped == 'EFD_CONTR':
            return None
        elif sped == 'ECF':
            return None
        elif sped == 'ECD':
            return None
    
    return None


def get_reg( row : list,  count_reg_rows : dict = None ) -> str:
    reg = row[1]
    if count_reg_rows is not None:
        if reg in count_reg_rows:
            count_reg_rows[reg] += 1
        else:
            count_reg_rows[reg] = 1

    return reg



def extract_fieldnames(output_path: str, chield_level: int = 3) -> bool:
    # Cria um workbook no modo padrão (não write_only)
    workbook = openpyxl.Workbook()
    workbook_reg = openpyxl.Workbook()

    # Define os tons de cinza
    fill_light_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_dark_gray = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    fill_blue = PatternFill(start_color="009DFF", end_color="009DFF", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, name='Calibri')


    for sped, version in [('ECF', '0010'), ('EFD_FISCAL', '019'), ('EFD_CONTR', '006'), ('ECD', '2024'), ('MANAD', '003')]:
        worksheet = workbook.create_sheet(title=f'{sped} - LEIAUTE_{version}')
        worksheet.append(['RELATÓRIO', 'REGISTRO', 'LAYOUT'])

        worksheet_reg = workbook_reg.create_sheet(title=f'{sped} - LEIAUTE_{version}')
        worksheet_reg.append(['REGISTRO', 'LAYOUT'])


       # Aplica o estilo de fundo azul e texto branco nos títulos
        for cell in worksheet[1]:
            cell.fill = fill_blue
            cell.font = white_font

        prefix = SPED_BY_VERSION[sped][version]['prefixo']

        # Obtém o dicionário
        sped_dict = load_sped_dictionary(sped, version, SPED_BY_VERSION[sped])
        if sped_dict is False:
            return False
        
        # ...
        reg_fieldnames = {}
        for reg in sped_dict:
            reg_fieldnames[reg] = {}
            for field in sped_dict[reg]['campos']:
                comp_field = f"{reg}{prefix}_{field}"
                if 'CNPJ' in field or 'CPF' in field:
                    reg_fieldnames[reg][comp_field] = 'STRING'
                elif 'DT_' in field and sped_dict[reg]['campos'][field]['tam'] == 8:
                    reg_fieldnames[reg][comp_field] = 'DATE'
                else:
                    if sped_dict[reg]['campos'][field]['tipo'] == 'text':
                        reg_fieldnames[reg][comp_field] = 'STRING' 
                    elif sped_dict[reg]['campos'][field]['tipo'] == 'numeric':
                        if 'VL' in field or 'VALOR' in field or 'QTD' in field or 'VLT' in field:
                            sped_dict[reg]['campos'][field]['tipo'] = 'NUMERIC'
                        else:
                            sped_dict[reg]['campos'][field]['tipo'] = 'STRING'
                    else:
                        sped_dict[reg]['campos'][field]['tipo'] = 'STRING'


        row_index = 0
        for report in reg_fieldnames:
            fill = fill_light_gray if row_index % 2 == 0 else fill_dark_gray
            regs = {}
            for field in reg_fieldnames[report]:
                reg = field[:4]
                if reg not in regs:
                    regs[reg] = []
                
                regs[reg].append(field)

            types = {}
            for field in reg_fieldnames[report]:
                reg = field[:4]
                if reg not in types:
                    types[reg] = []
                
                types[reg].append(reg_fieldnames[report][field])

            for reg in regs:
                row_data = [reg] + list(regs[reg])
                worksheet_reg.append(row_data)

                # Acessa a última linha adicionada com base em max_row
                last_row = worksheet_reg.max_row
                                
                # Aplica o estilo de preenchimento alternado
                for cell in worksheet_reg[last_row]:  # Itera sobre as células da última linha
                    cell.fill = fill

                row_data = [reg] + list(types[reg])
                worksheet_reg.append(row_data)

                # Acessa a última linha adicionada com base em max_row
                last_row = worksheet_reg.max_row
                                
                # Aplica o estilo de preenchimento alternado
                for cell in worksheet_reg[last_row]:  # Itera sobre as células da última linha
                    cell.fill = fill

            row_index += 2



        # Obtém os nomes dos campos do relatório
        report_fieldnames = {}
        report_fieldnames_types = {}
        report_reg_compl = deepcopy(REPORT_REG_COMPL)
        compl_tables = load_compl_tables()
        report_fieldnames, report_fieldnames_types = get_report_fieldnames( sped, prefix, sped_dict, chield_level, report_fieldnames, report_reg_compl, compl_tables, report_fieldnames_types )

        row_index = 0
        for report in report_fieldnames:
            fill = fill_light_gray if row_index % 2 == 0 else fill_dark_gray
            regs = {}
            for field in report_fieldnames[report]:
                reg = field[:4]
                if reg not in regs:
                    regs[reg] = []
                
                regs[reg].append(field)

            types = {}
            for field in report_fieldnames_types[report]:
                reg = field[:4]
                if reg not in types:
                    types[reg] = []
                
                types[reg].append(report_fieldnames_types[report][field])

            for reg in regs:
                row_data = [report, reg] + list(regs[reg])
                worksheet.append(row_data)

                # Acessa a última linha adicionada com base em max_row
                last_row = worksheet.max_row
                                
                # Aplica o estilo de preenchimento alternado
                for cell in worksheet[last_row]:  # Itera sobre as células da última linha
                    cell.fill = fill

                row_data = [report, reg] + list(types[reg])
                worksheet.append(row_data)

                # Acessa a última linha adicionada com base em max_row
                last_row = worksheet.max_row
                                
                # Aplica o estilo de preenchimento alternado
                for cell in worksheet[last_row]:  # Itera sobre as células da última linha
                    cell.fill = fill

            row_index += 2

    # Remove a planilha padrão criada automaticamente
    if "Sheet" in workbook.sheetnames:
        std = workbook["Sheet"]
        workbook.remove(std)

    # Remove a planilha padrão criada automaticamente
    if "Sheet" in workbook_reg.sheetnames:
        std = workbook_reg["Sheet"]
        workbook_reg.remove(std)

    # Salva o arquivo
    workbook.save( os.path.join(output_path, 'FIELDS_BY_REPORT.xlsx') )

    workbook_reg.save( os.path.join(output_path, 'FIELDS_BY_REG.xlsx') )

    return True




def custom_json_dumps(obj, indent=2, level=0, open_levels=2):
    """
    Recursively serialize obj to a JSON string.
    Only the first `open_levels` levels will be pretty printed;
    deeper levels are rendered in a compact form.
    
    Parameters:
      obj: The object (dict, list, etc.) to serialize.
      indent: Number of spaces for each indent level.
      level: The current recursive level (internal use).
      open_levels: Levels that will be pretty printed.
    
    Returns:
      A JSON string.
    """
    if isinstance(obj, dict):
        if level < open_levels:
            items = []
            for key, value in obj.items():
                # Serialize the key using json.dumps to handle quotes, etc.
                serialized_key = json.dumps(key, ensure_ascii=False)
                # Recursively serialize the value
                serialized_value = custom_json_dumps(value, indent, level + 1, open_levels)
                items.append(' ' * ((level + 1) * indent) + f"{serialized_key} : {serialized_value}")
            # Join items with newlines and proper indentation for the braces
            return '{\n' + ',\n'.join(items) + '\n' + ' ' * (level * indent) + '}'
        else:
            # For levels deeper than open_levels, use a compact representation
            return json.dumps(obj, separators=(',', ':'), ensure_ascii=False)
        
    elif isinstance(obj, list):
        if level < open_levels:
            items = []
            for item in obj:
                serialized_item = custom_json_dumps(item, indent, level + 1, open_levels)
                items.append(' ' * ((level + 1) * indent) + serialized_item)
            return '[\n' + ',\n'.join(items) + '\n' + ' ' * (level * indent) + ']'
        else:
            return json.dumps(obj, separators=(',', ':'))
    else:
        # For simple types, just use json.dumps.
        return json.dumps(obj)

def dict_to_json_file(data, file_path, indent=2, open_levels=2):
    """
    Converts a Python dictionary into a JSON file with only the first
    `open_levels` levels pretty printed. Deeper levels are rendered compactly.
    
    Parameters:
      data: The Python dictionary to convert.
      file_path: The path to the output JSON file.
      indent: Number of spaces to use for indentation.
      open_levels: Number of levels to pretty print.
    """
    json_str = custom_json_dumps(data, indent=indent, level=0, open_levels=open_levels)
    with open(file_path, 'w', encoding='UTF-8') as f:
        f.write(json_str)



def extract_fieldnames_v2(output_path: str, chield_level: int = 3) -> bool:
    table_schames = {}

    for sped in SPED_BY_VERSION:
        if sped == 'MANAD':
            continue
        for version in SPED_BY_VERSION[sped]:
            prefix = SPED_BY_VERSION[sped][version]['prefixo']

            # Obtém o dicionário
            sped_dict = load_sped_dictionary(sped, version, SPED_BY_VERSION[sped])
            if sped_dict is False:
                return False
            
            reg_fieldnames = {}
            for reg in sped_dict:
                reg_fieldnames[reg] = {}
                for field in sped_dict[reg]['campos']:
                    field_type = 'STRING'
                    _field = field.replace('/', '_').replace('\n', '').replace(";", "")
                    if 'VL' in _field or \
                        'VLR' in _field or \
                        'VAL_' in _field or \
                        'VALOR' in _field or \
                        'QTD' in _field or \
                        'QUANT' in _field:

                        field_type = 'NUMERIC'
                        if 'IND_' in _field or 'ESPERADO' in _field:
                            field_type = 'STRING'

                    
                    comp_field = f"{reg}{prefix}_{_field}"
                    reg_fieldnames[reg][comp_field] = field_type

            # REG_REPORT
            for reg_report in reg_fieldnames:
                if reg_report.startswith('9') or reg_report.endswith('990'):
                    continue
                table_name = f"{sped}_{reg_report}"
                if table_name not in table_schames:
                    table_schames[table_name] = {}

                for field in reg_fieldnames[reg_report]:
                    _field = field.replace('/', '_').replace('\n', '').replace(";", "")
                    if _field in table_schames[table_name]:
                        continue
                    field_type = reg_fieldnames[reg_report][field]
                    table_schames[table_name][_field] = {'type' : field_type, 'mode' : 'NULLABLE', 'treat_type' : field_type }

            # Obtém os nomes dos campos do relatório
            report_fieldnames = {}
            # report_fieldnames_types = {}
            report_reg_compl = deepcopy(REPORT_REG_COMPL)
            compl_tables = load_compl_tables()
            report_fieldnames = get_report_fieldnames( sped, prefix, sped_dict, chield_level, report_fieldnames, report_reg_compl, compl_tables )

            for report in report_fieldnames:
                table_name = f"Relatorio_{sped}_{report}"
                if table_name not in table_schames:
                    table_schames[table_name] = {}

                for field in report_fieldnames[report]:
                    _field = field.replace('/', '_').replace('\n', '').replace(";", "")
                    if _field in table_schames[table_name]:
                        continue
                    
                    field_type = 'STRING'
                    if 'VL' in _field or \
                        'VLR' in _field or \
                        'VAL_' in _field or \
                        'VALOR' in _field or \
                        'QTD' in _field or \
                        'QUANT' in _field:

                        field_type = 'NUMERIC'
                        if 'IND_' in _field or 'ESPERADO' in _field or 'ALIQ' in _field:
                            field_type = 'STRING'

                    table_schames[table_name][_field] = {
                        'type' : field_type, 
                        'mode' : 'NULLABLE', 
                        'treat_type' : field_type 
                    }

    # ENRICH FIELDS
    for table_name in table_schames:
        table_schame = None
        if '0000' not in table_name:
            table_schame = {
                'CompetenciaArquivo' : {'type' : 'STRING', 'mode' : 'REQUIRED', 'treat_type' : 'STRING'},
                'CNPJ' : {'type' : 'STRING', 'mode' : 'REQUIRED', 'treat_type' : 'STRING'},
                **dict(sorted(table_schames[table_name].items(), key=lambda x : x[0])),
                'HashFile' : {'type' : 'STRING', 'mode' : 'REQUIRED', 'treat_type' : 'STRING'},
            }
        else:
            table_schame = {
                'CompetenciaArquivo' : {'type' : 'STRING', 'mode' : 'REQUIRED', 'treat_type' : 'STRING'},
                'CNPJ' : {'type' : 'STRING', 'mode' : 'REQUIRED', 'treat_type' : 'STRING'},
                **dict(sorted(table_schames[table_name].items(), key=lambda x : x[0])),
                'HashFile' : {'type' : 'STRING', 'mode' : 'REQUIRED', 'treat_type' : 'STRING'},
                'DataTransmissao' : {'type' : 'DATETIME', 'mode' : 'NULLABLE', 'treat_type' : 'DATETIME'},
                'ProcTime' : {'type' : 'DATETIME', 'mode' : 'NULLABLE', 'treat_type' : 'DATETIME'},
                'FileName' : {'type' : 'STRING', 'mode' : 'NULLABLE', 'treat_type' : 'FILENAME'}
            }
        table_schames[table_name] = table_schame
        # # table_schames[table_name]['CNPJ'] = {'type' : 'STRING', 'mode' : 'REQUIRED', 'treat_type' : 'STRING'}
        # # table_schames[table_name]['CompetenciaArquivo'] = {'type' : 'DATE', 'mode' : 'REQUIRED', 'treat_type' : 'DATE'}
        # # table_schames[table_name]['HashFile'] = {'type' : 'STRING', 'mode' : 'REQUIRED', 'treat_type' : 'STRING'}
        # # table_schames[table_name]['FileName'] = {'type' : 'STRING', 'mode' : 'NULLABLE', 'treat_type' : 'FILENAME'}
        # for field in table_schames[table_name]:
        #     table_schames[table_name][field]

    schame_analysis = {}
    for table_name in table_schames:
        for field in table_schames[table_name]:
            base_field = field.split('_', 1)[-1]
            if base_field not in schame_analysis:
                schame_analysis[base_field] = {}
            field_type = table_schames[table_name][field]['type']
            if field_type not in schame_analysis[base_field]:
                schame_analysis[base_field][field_type] = []
            if table_name not in schame_analysis[base_field][field_type]:
                schame_analysis[base_field][field_type].append(table_name)

    for base_field in list(schame_analysis):
        # if not base_field.endswith('_'):
        #     schame_analysis.pop(base_field, None)

        # if 'DATE' not in schame_analysis[base_field]:
        #     schame_analysis.pop(base_field, None)

        if len(schame_analysis[base_field]) == 1:
            schame_analysis.pop(base_field, None)

    dict_to_json_file(data=table_schames, file_path=os.path.join(output_path, 'tables_schemes.json'))

    return True



def get_file_info( row : list, file_info : dict, sped : str, version : str, sped_dict : dict, sped_by_version : dict ) -> bool:
    reg_0000_info = sped_dict['0000']

    # CNPJ
    file_info['CNPJ'] = row[reg_0000_info['campos']['CNPJ']['ordem']]
    file_info['Nome'] = row[reg_0000_info['campos']['NOME']['ordem']]
    
    if file_info['Tipo'] == 'ECD':
        # LEIAUTE
        file_info['Leiaute'] = sped_by_version[sped][version]["version"]

        # Competencia
        file_info['Competencia'] = row[reg_0000_info['campos']['DT_FIN']['ordem']][-4:]

        # Retificador ou Original
        if 'IND_FIN_ESC' not in reg_0000_info['campos']:
            file_info['Retificador ou Original'] = '-1'
        else:
            if len(row) > reg_0000_info['campos']['IND_FIN_ESC']['ordem']:
                file_info['Retificador ou Original'] = row[reg_0000_info['campos']['IND_FIN_ESC']['ordem']]
            else:
                file_info['Retificador ou Original'] = '-1'

        # Unico
        if file_info['Retificador ou Original'] == '1':
            file_info['Retificador ou Original'] = 'Retificador'
        elif file_info['Retificador ou Original'] == '0':
            file_info['Retificador ou Original'] = 'Original'
        elif file_info['Retificador ou Original'] == '-1':
            file_info['Retificador ou Original'] = 'Unico'
    

    elif file_info['Tipo'] == 'ECF':
        # LEIAUTE
        file_info['Leiaute'] = version
        
        # Competencia        
        competencia = row[reg_0000_info['campos']['DT_FIN']['ordem']][-4:]
        file_info['Competencia'] = competencia
                
        # Retificador ou Original
        file_info['Retificador ou Original'] = row[reg_0000_info['campos']['RETIFICADORA']['ordem']]
        if file_info['Retificador ou Original'] == 'S':
            file_info['Retificador ou Original'] = 'Retificador'
        elif file_info['Retificador ou Original'] in ('N', 'F'):
            file_info['Retificador ou Original'] = 'Original'


    elif file_info['Tipo'] == 'EFD_FISCAL':
        file_info['IE'] = row[reg_0000_info['campos']['IE']['ordem']]

        # LEIAUTE
        file_info['Leiaute'] = version

        # Competencia
        competencia = row[reg_0000_info['campos']['DT_FIN']['ordem']][-6:]
        file_info['Competencia'] = f'{competencia[2:]}-{competencia[:2]}'
                
        # Retificador ou Original
        file_info['Retificador ou Original'] = row[reg_0000_info['campos']['COD_FIN']['ordem']]
        if file_info['Retificador ou Original'] == '1':
            file_info['Retificador ou Original'] = 'Retificador'
        elif file_info['Retificador ou Original'] == '0':
            file_info['Retificador ou Original'] = 'Original'


    elif file_info['Tipo'] == 'EFD_CONTR':
        # LEIAUTE
        file_info['Leiaute'] = version

        # Competencia
        competencia = row[reg_0000_info['campos']['DT_FIN']['ordem']][-6:]
        file_info['Competencia'] = f'{competencia[2:]}-{competencia[:2]}'
                
        # Retificador ou Original
        file_info['Retificador ou Original'] = row[reg_0000_info['campos']['TIPO_ESCRIT']['ordem']]
        if file_info['Retificador ou Original'] == '1':
            file_info['Retificador ou Original'] = 'Retificador'
        elif file_info['Retificador ou Original'] == '0':
            file_info['Retificador ou Original'] = 'Original'


    elif file_info['Tipo'] == 'MANAD':
        # LEIAUTE
        file_info['Leiaute'] = version

        # Competencia
        competencia_ini = row[reg_0000_info['campos']['DT_INI']['ordem']][-6:]
        competencia_fin = row[reg_0000_info['campos']['DT_FIN']['ordem']][-6:]
        file_info['Competencia'] = f"{competencia_ini[2:]}-{competencia_ini[:2]} - {competencia_fin[2:]}-{competencia_fin[:2]}"
    
    return True


def write_reg_row( reg : str, reg_dict : dict, file_info : dict, reg_file_pointers : dict ) -> bool:

    row = {
        'CompetenciaArquivo' : file_info['Competencia'],
        'CNPJ' : file_info['CNPJ'],
        **reg_dict,
        'HashFile' : reg_file_pointers[reg]['hash_file'], 
        'FileName' : reg_file_pointers[reg]['file_name']
        }
    
    # reg_file_pointers[reg]['data'].append(row)
    reg_file_pointers[reg]['writer'].writerow(row)
    reg_file_pointers[reg]['num_rows'] += 1

    return True


def check_empty_file(sped : str, reg_file_info : dict) -> bool:
    blocos = None
    if sped == 'EFD_FISCAL':
        blocos = ['C', 'D', 'G']
    elif sped == 'EFD_CONTR':
        blocos = ['A', 'C', 'D', 'F']
    else:
        return None

    for reg, reg_info in reg_file_info.items():
        num_rows = reg_info.get('num_rows', 0)
        if num_rows == 0:
            continue
        
        if any(reg.startswith(bloco) for bloco in blocos):
            if not reg.endswith(('001', '990')):
                return False
    
    return True


def generate_reports_sped(
        file_path : str,
        output_path : str,
        preprocessing_folder : str = None,
        files_considered : dict = {},
        map_reports : dict = {},
        chield_level : int = 3,
        delimiter : str = DELIMITER,
        dt_proc : dict = {},
        parquet_files = False,
        overwrite = False
        ) -> bool:
    logs = []
    
    # CREATE_FOLDERS
    reports_folder = os.path.join(output_path, 'reports')
    regs_folder = os.path.join(output_path, 'regs')
    metadata_folder = os.path.join(output_path, 'metadata')
    os.makedirs( reports_folder, exist_ok=True )
    os.makedirs( regs_folder, exist_ok=True )
    os.makedirs( metadata_folder, exist_ok=True )


    file_info = {
        'Tipo' : '', # OK
        'Leiaute' : '', # OK
        'CNPJ' : '',
        'Nome' : '',
        'IE' : '',
        'Competencia' : '',
        'Retificador ou Original' : '',
        'DT_INI' : '',
        'DT_FIN' : '',
        'DT_TRANS' : '',
        'DT_ASSINATURA' : '',
        'REG0000' : None,
        'Hashfile' : '', # OK
        'Filename' : file_path,
        'file_size' : '', # OK
        'line_qty' : '',
        'empty_file' : '',
        'Encoding' : '', # OK
        'Processado' : None,
        'Repetido' : False,
        'Termino sem registro 9999' : False,
        'Layout com erro' : False,
        'Estrutura com erro' : False,
        'Linha fora do padrao SPED' : False,
    }

    father_by_reg = { '' : {} }
    allow_father_by_reg = { '' : True }
    son_by_father = { 'position' : 0, 'sons' : {} }
    report_fieldnames = {}
    count_reg_rows = {}
    report_reg_compl = deepcopy( REPORT_REG_COMPL )
    compl_tables = load_compl_tables()
    compl_info = {}
    file_ended_with_9999 = None

    if os.path.exists( file_path ) is False:
        file_info['Processado'] = False
        return file_info, logs

    # File size in megabytes
    file_info['file_size'] = os.path.getsize( file_path ) / 1048576

    # Open '.txt' file
    detected_encoding, score = detect_encoding( file_path )
    file_info['Encoding'] = detected_encoding
    
    file_in = open( file=file_path, mode='r', newline='', encoding=detected_encoding, errors='replace' )
    # reader = csv.reader( f_in, delimiter=delimiter, quotechar = None )
    try:
        line = next( file_in )
    except:
        log_message = 'Não foi possível ler a primeira linha do arquivo.'
        log_row = {'function' : 'generate_reports_sped', 'message' : log_message}
        logs.append(log_row)

        file_info['Processado'] = False
        return file_info, logs
    

    # GET SPED INFO
    version, sped, prefix, delimiter_start = get_sped_info( line.split(delimiter), SPED_BY_VERSION, logs, file_path, detected_encoding, delimiter )
    if version is False:
        log_message = f'Versão do arquivo não identificada.'
        log_row = { 'function' : 'generate_reports_sped', 'message' : log_message }
        logs.append(log_row )

        file_info['Processado'] = False
        return file_info, logs
    file_info['Tipo'] = sped

    summary_by_report = deepcopy( SUMMARY_BY_REPORT[sped] )
    summary_by_field = deepcopy( SUMMARY_BY_FIELD[sped] )
    
    # Get dictionary
    sped_dict = load_sped_dictionary( sped, version, SPED_BY_VERSION[sped] )
    if sped_dict is False:
        log_message = f'Manual sped não cadastrado no sistema.'
        log_row = {
            'function' : 'generate_reports_sped',
            'message' : log_message
            }
        logs.append(log_row)
        file_info['Processado'] = False
        return file_info, logs


    # BUILD THE FIRST ROW
    row = build_row(
        line=line, 
        delimiter=delimiter, 
        delimiter_start=delimiter_start, 
        file=file_in, 
        row_id=1,
        file_info=file_info,
        sped_dict=sped_dict,
        logs=logs
        )
    if row == []:
        return False, logs


    get_file_info( row, file_info, sped, version, sped_dict, SPED_BY_VERSION )
    hash_file, signdatetime = generate_hash_and_signdate( file_path )
    file_info['Hashfile'] = hash_file
    file_info['DT_ASSINATURA'] = signdatetime
    file_info['DT_TRANS'] = get_dt_trans(hash_file, dt_proc, os.path.basename(file_path))

    if sped not in files_considered:
        files_considered[sped] = {}

    if hash_file in files_considered[sped]:
        file_info['Processado'] = None
        file_info['Repetido'] = True
        return file_info, logs


    # Build first reg info
    reg = get_reg( row, count_reg_rows )
    cnpj = file_info['CNPJ']
    retificador = file_info['Retificador ou Original']


    # Get report Field Names
    # try:
    report_fieldnames = get_report_fieldnames(
        sped, 
        prefix, 
        sped_dict, 
        chield_level, 
        report_fieldnames, 
        report_reg_compl, 
        compl_tables
    )

    
    ret_report_fieldnames = get_reg_report_fieldnames(
        prefix,
        sped_dict,
        compl_tables
        )

    layout_error = False
    reg_dict, row_error = build_row_dict( reg, row, sped_dict, prefix, 1, compl_tables, summary_by_field, logs )
    if row_error:
        layout_error = True

    # GET DT_INI, DT_FIN
    for field in reg_dict:
        if reg_dict[field] == '':
            continue
        if field.endswith('DT_INI'):
            file_info['DT_INI'] = f"{reg_dict[field][4:]}-{reg_dict[field][2:4]}-{reg_dict[field][:2]}"
        if field.endswith('DT_FIN'):
            file_info['DT_FIN'] = f"{reg_dict[field][4:]}-{reg_dict[field][2:4]}-{reg_dict[field][:2]}"
    file_info['REG0000'] = copy(reg_dict)
    

    # CHECK IF PRE-PROCESSED FILE EXISTS
    if preprocessing_folder:
        pfile_name = f"{file_info['Hashfile']}.pickle"
        pickle_folder = os.path.join(preprocessing_folder, 'SPED_REPORTS', 'PICKLES')
        os.makedirs(pickle_folder, exist_ok=True)
        pfile_path = os.path.join( pickle_folder, pfile_name )

        if os.path.exists(pfile_path) is True:
            f_in = open(pfile_path, mode='rb')
            pfile_info = pickle.load(f_in)
            f_in.close()

            if pfile_info['file_info']['Retificador ou Original'] != file_info['Retificador ou Original']:
                pfile_info['file_info']['Retificador ou Original'] = file_info['Retificador ou Original']
                with open(pfile_path, mode='wb') as pf_out:
                    pickle.dump( pfile_info, pf_out )

            # UPDATE THE 'empty_file' FIELD
            empty_file = check_empty_file(sped=sped, reg_file_info=pfile_info['reg_info'])
            if 'empty_file' not in pfile_info['file_info'] or pfile_info['file_info']['empty_file'] != empty_file:
                pfile_info['file_info']['empty_file'] = empty_file
                with open(pfile_path, mode='wb') as pf_out:
                    pickle.dump( pfile_info, pf_out )


            # UPDATE DT_TRANS
            if file_info['DT_TRANS'] != '' and pfile_info['file_info'].get('DT_TRANS', '') == '':
                pfile_info['file_info']['DT_TRANS'] = file_info['DT_TRANS']
                with open(pfile_path, mode='wb') as pf_out:
                    pickle.dump( pfile_info, pf_out )
            if pfile_info['file_info']['DT_TRANS'].isnumeric():
                dt_trans = pfile_info['file_info']['DT_TRANS']
                dt_trans = datetime.strptime(dt_trans, "%Y%m%d%H%M%S").strftime('%Y-%m-%dT%H:%M:%S')
                pfile_info['file_info']['DT_TRANS'] = dt_trans
                with open(pfile_path, mode='wb') as pf_out:
                    pickle.dump( pfile_info, pf_out )


            # TXT FOLDER
            txt_folder = os.path.join( preprocessing_folder, 'SPED_REPORTS', 'TXTS', pfile_info['file_info']['CNPJ'][:8], pfile_info['file_info']['CNPJ'], pfile_info['file_info']['Tipo'], pfile_info['file_info']['Competencia'])
            txt_path = os.path.join( txt_folder, pfile_info['txt_name'])

            # Identificar possíveis motivos
            if 'DT_TRANS' not in pfile_info['file_info'] or \
                pfile_info['file_info']['DT_TRANS'] in ('', None) or \
                'txt_name' not in pfile_info or \
                ('DT_TRANS_-' in pfile_info['txt_name'] and pfile_info['file_info'].get('DT_TRANS', '') != '') or \
                os.path.exists( txt_path ) is False:

                # TXT FOLDER
                txt_folder = os.path.join(preprocessing_folder, 'SPED_REPORTS', 'TXTS')
                os.makedirs(txt_folder, exist_ok=True)

                # FILE
                cnpj = pfile_info['file_info']['CNPJ']
                sped = pfile_info['file_info']['Tipo']
                version = pfile_info['file_info']['Leiaute']
                competencia = pfile_info['file_info']['Competencia']
                retificador = pfile_info['file_info']['Retificador ou Original']
                if pfile_info['file_info']['DT_TRANS'] not in (None, ''):
                    datatrans = pfile_info['file_info']['DT_TRANS'].replace('-', '').replace(':', '').replace('T', '').replace(' ', '')
                else:
                    datatrans = ''
                hashfile = file_info['Hashfile']

                # TXT NAME
                txt_name = f"CNPJ_{cnpj}-SPED_{sped}-VERSION_{version}-COMPETENCIA_{competencia}-RETIFICADOR_{retificador}-DT_TRANS_{datatrans}-HASH_{hashfile}.zip"

                # TXT SUB FOLDER
                txt_sub_folder = os.path.join( txt_folder, cnpj[:8], cnpj, sped, competencia )
                os.makedirs( txt_sub_folder, exist_ok=True )

                # ZIP PATH
                ziptxt_path = os.path.join(txt_sub_folder, txt_name)


                files = { os.path.basename(file_path) : txt_name.replace('.zip', '.txt') }
                input_folder = os.path.dirname(file_path)
                io_tools.move_files_to_zip( input_folder=input_folder, files=files, zip_output_path=ziptxt_path )


                # Update pfile
                pfile_info['txt_name'] = txt_name
                # UPDATE PFILE
                with open(pfile_path, mode='wb') as pf_out:
                    pickle.dump( pfile_info, pf_out )


            # COPY PFILE TO LOCAL FOLDER
            pfile_path_to = os.path.join(metadata_folder, pfile_name)
            shutil.copy(pfile_path, pfile_path_to)
            
            return pfile_info['file_info'], pfile_info['logs']


    if sped not in map_reports:
        map_reports[sped] = {
            'CNPJ_competencia' : [(file_info['CNPJ'], file_info['Competencia'])],
            'reports' : list(report_fieldnames)
        }
    else:
        cnpj_competencia = (file_info['CNPJ'], file_info['Competencia'])
        if cnpj_competencia not in map_reports[sped]['CNPJ_competencia']:
            map_reports[sped]['CNPJ_competencia'].append(cnpj_competencia)


    # File pointers
    file_pointers = build_pointers( 
        sped = sped, 
        version = version, 
        cnpj = cnpj, 
        competencia  =file_info['Competencia'], 
        retificador_original = file_info['Retificador ou Original'], 
        sped_dict  =sped_dict, 
        output_path = reports_folder, 
        report_fieldnames = report_fieldnames, 
        hash_file = hash_file, 
        file_name = os.path.basename(file_path)
        )

    reg_file_pointers = build_reg_pointers( 
        sped = sped, 
        version = version, 
        cnpj = cnpj, 
        competencia = file_info['Competencia'], 
        retificador_original = file_info['Retificador ou Original'], 
        prefix = prefix,
        sped_dict  = sped_dict, 
        output_path = regs_folder, 
        ret_report_fieldnames = ret_report_fieldnames,
        hash_file = hash_file, 
        file_name = os.path.basename(file_path)
        )
    

    write_reg_row( reg, reg_dict, file_info, reg_file_pointers )

    father_by_reg[reg] = reg_dict
    allow_father_by_reg[reg] = True if reg_dict != {} else False
    

    for compl_reg in report_reg_compl[sped]:
        compl_info[compl_reg] = {}

    
    try:
        # Iterate throught each reg info
        row_id = 1
        for line in file_in:
            row_id += 1

            # BUILD ROW
            row = build_row(
                line=line, 
                delimiter=delimiter, 
                delimiter_start=delimiter_start, 
                file=file_in, 
                row_id=row_id,
                file_info=file_info,
                sped_dict=sped_dict,
                logs=logs
                )
            if row == []:
                file_info['Fora do padrao SPED'] = True
                layout_error = True
                break

            reg = get_reg( row, count_reg_rows )
            

            reg_dict, row_error = build_row_dict( reg, row, sped_dict, prefix, row_id, compl_tables, summary_by_field, logs )
            if row_error:
                layout_error = True
                continue

            write_reg_row( reg, reg_dict, file_info, reg_file_pointers )
            reg_father = sped_dict[reg]['pai']


            # if reg not in father_by_reg or reg not in _report_reg[sped]:

            # DESENV 20230913
            reg_ancestors_list = get_parents( reg, sped_dict, chield_level, REPORT_REG[sped] )


            for reg_ancestrals in reg_ancestors_list:
                _son_by_father = son_by_father

                # Get data position
                for _reg_son in reg_ancestrals:
                    if _reg_son not in _son_by_father['sons']:
                        _son_by_father['sons'][_reg_son] = {
                            'rows' : [],
                            'size' : 1,
                            'positions' : [],
                            'sons' : {},
                        }
                    _son_by_father = _son_by_father['sons'][_reg_son]
                    
                if reg not in _son_by_father['sons']:
                    _son_by_father['sons'][reg] = {
                        'rows' : [reg_dict],
                        'size' : 1,
                        'positions' : [_son_by_father['size'] - 1],
                        'sons' : {}
                    }

                else:
                    _son_by_father['sons'][reg]['rows'].append(reg_dict)
                    _son_by_father['sons'][reg]['positions'].append(_son_by_father['size'] - 1)
                    _son_by_father['sons'][reg]['size'] += 1


            if reg in father_by_reg and reg in REPORT_REG[sped]:
                write_row(
                    reg,
                    file_pointers,
                    report_reg_compl[sped],
                    son_by_father,
                    father_by_reg,
                    allow_father_by_reg,
                    compl_info,
                    summary_by_report,
                    file_info
                    )

            # Update REG INFO
            allow_father_by_reg[reg] = True if reg_dict != {} and allow_father_by_reg[reg_father] is True else False
            father_by_reg[reg] = { **father_by_reg[reg_father], **reg_dict }


            # compl info
            if reg in compl_info:
                key_value = get_compl_key( reg, report_reg_compl[sped][reg], father_by_reg[reg] )
                if key_value not in compl_info[reg]:
                    compl_info[reg][key_value] = father_by_reg[reg]
                else:
                    pass

            if reg == '9999' or reg == '9990':
                file_ended_with_9999 = True
                break

        file_in.close()
        
    except Exception as e:
        log_message = f'Há erro na estrutura do arquivo. {e}'
        log_row = {'function' : 'generate_reports_sped', 'message' : log_message}
        logs.append(log_row)

        file_info['Processado'] = False
        file_info['Estrutura com erro'] = True

    if file_ended_with_9999 is not True and layout_error is not True:
        file_ended_with_9999 = False

    if file_ended_with_9999 is False:
        log_message = f'Arquivo fora do padrão sped, não terminou com o registro 9999.'
        log_row = {'function' : 'generate_reports_sped', 'message' : log_message, 'error_type' : 'file_error'}
        logs.append(log_row)

        file_info['Processado'] = False
        file_info['Termino sem registro 9999'] = True

    
    if file_info['Linha fora do padrao SPED'] is True:
        log_message = f'Arquivo com linha fora do padrão sped.'
        log_row = {'function' : 'generate_reports_sped', 'message' : log_message, 'error_type' : 'file_error'}
        logs.append(log_row)
        file_info['Processado'] = False

    

    # Write Remaining
    if layout_error is not True:
        for reg in father_by_reg:
            write_row(
                reg, 
                file_pointers, 
                report_reg_compl[sped], 
                son_by_father, 
                father_by_reg, 
                allow_father_by_reg, 
                compl_info, 
                summary_by_report, 
                file_info 
                )
    
    if  layout_error is True:
        file_info['Processado'] = False
        file_info['Layout com erro'] = True

    # Close files
    list_of_reports = list(file_pointers)
    for reg in list_of_reports:
        if file_info['Processado'] is False:
            file_pointers[reg]['num_rows'] = 0
            file_pointers[reg]['pointer'].truncate(0)
        file_pointers[reg]['pointer'].close()
        file_pointers[reg]['pointer'] = None
        file_pointers[reg]['writer'] = None

        if file_pointers[reg]['num_rows'] == 0:
            if os.path.exists(file_pointers[reg]['file_path']):
                os.remove(file_pointers[reg]['file_path'])
            file_pointers.pop(reg)
        else:
            if parquet_files is True:
                convert_csv2parquet(file_pointers[reg]['file_path'], overwrite=overwrite)

            file_pointers[reg].pop('pointer')
            file_pointers[reg].pop('writer')
            file_pointers[reg]['file_name'] = os.path.basename(file_pointers[reg]['file_path'])
            file_pointers[reg].pop('file_path')

    # FILL 'empty_file' field
    file_info['empty_file'] = check_empty_file(sped=sped, reg_file_info=reg_file_pointers)

    # Close reg files
    list_of_reports = list(reg_file_pointers)
    for reg in list_of_reports:
        if file_info['Processado'] is False:
            reg_file_pointers[reg]['num_rows'] = 0
            reg_file_pointers[reg]['pointer'].truncate(0)
        reg_file_pointers[reg]['pointer'].close()
        reg_file_pointers[reg]['pointer'] is None
        reg_file_pointers[reg]['writer'] is None

        if reg_file_pointers[reg]['num_rows'] == 0:
            if os.path.exists(reg_file_pointers[reg]['file_path']):
                os.remove(reg_file_pointers[reg]['file_path'])
            reg_file_pointers.pop(reg)
        else:
            if parquet_files is True:
                convert_csv2parquet(reg_file_pointers[reg]['file_path'], overwrite=overwrite)

            reg_file_pointers[reg].pop('pointer')
            reg_file_pointers[reg].pop('writer')
            reg_file_pointers[reg]['file_name'] = os.path.basename(reg_file_pointers[reg]['file_path'])
            reg_file_pointers[reg].pop('file_path')



    # Validation
    for report in summary_by_report:
        for field in summary_by_report[report]:
            if summary_by_report[report][field] != summary_by_field[field]:
                log_message = f"Erro de validação (report_{report} field_{field} - TXT[{summary_by_field[field][0]}, {str(summary_by_field[field][1])}] - REPORT[{summary_by_report[report][field][0]}, {str(summary_by_report[report][field][1])}] )"
                log_row = {'function' : 'generate_reports_sped', 'message' : log_message, 'error_type' : 'validation_error'}
                logs.append(log_row)


    file_info['line_qty'] = row_id
    if file_info['Processado'] is None:
        file_info['Processado'] = True


    # WRITE PFILE_INFO
    if file_info['Processado'] is True and preprocessing_folder is not None:
        ################## TXT ####################
        # COPY TXT TO PREPROCESS_FOLDER
        txt_folder = os.path.join(preprocessing_folder, 'SPED_REPORTS', 'TXTS')
        os.makedirs(txt_folder, exist_ok=True)
        cnpj = file_info['CNPJ']
        sped = file_info['Tipo']
        version = file_info['Leiaute']
        competencia = file_info['Competencia']
        retificador = file_info['Retificador ou Original']
        if file_info['DT_TRANS'] not in (None, ''):
            datatrans = file_info['DT_TRANS'].replace('-', '').replace(':', '').replace('T', '').replace(' ', '')
        else:
            datatrans = ''
        hashfile = file_info['Hashfile']
        txt_name = f"CNPJ_{cnpj}-SPED_{sped}-VERSION_{version}-COMPETENCIA_{competencia}-RETIFICADOR_{retificador}-DT_TRANS_{datatrans}-HASH_{hashfile}.zip"
        txt_sub_folder = os.path.join( txt_folder, cnpj[:8], cnpj, sped, competencia )
        os.makedirs( txt_sub_folder, exist_ok=True )
        ziptxt_path = os.path.join(txt_sub_folder, txt_name)

        files = { os.path.basename(file_path) : txt_name.replace('.zip', '.txt') }
        input_folder = os.path.dirname(file_path)
        io_tools.move_files_to_zip( input_folder=input_folder, files=files, zip_output_path=ziptxt_path )

        ############### PICKLE #################3
        pfile_name = f"{file_info['Hashfile']}.pickle"
        pfile_path = os.path.join( metadata_folder, pfile_name )
        with open( pfile_path, mode='wb' ) as pf_out:
            pfile_info = {
                "file_info" : copy(file_info),
                "report_info" : copy(file_pointers),
                "reg_info" : copy(reg_file_pointers),
                "txt_name" : txt_name,
                "logs" : copy(logs),
            }
            pickle.dump(pfile_info, pf_out)

        ############### IND_REPORTS ###################
        # MOVE REPORTS TO PREPROCESS_FOLDER
        ind_reports_folder = os.path.join( preprocessing_folder, 'SPED_REPORTS', 'IND_REPORTS' )
        zipfile_path = os.path.join( ind_reports_folder, f"REPORTS-HASH_{file_info['Hashfile']}.zip" )
        list_of_files = [file_pointers[reg]['file_name'] for reg in file_pointers]
        io_tools.move_files_to_zip( input_folder=reports_folder, files=list_of_files, zip_output_path=zipfile_path )
        for file_name in list_of_files:
            rfile_path = os.path.join(reports_folder, file_name)
            os.remove(rfile_path)
        
        ############### IND_REGS ###################
        # MOVE REPORTS TO PREPROCESS_FOLDER
        ind_regs_folder = os.path.join( preprocessing_folder, 'SPED_REPORTS', 'IND_REGS' )
        zipfile_path = os.path.join( ind_regs_folder, f"REGS-HASH_{file_info['Hashfile']}.zip" )
        list_of_files = [reg_file_pointers[reg]['file_name'] for reg in reg_file_pointers]
        io_tools.move_files_to_zip( input_folder=regs_folder, files=list_of_files, zip_output_path=zipfile_path )
        for file_name in list_of_files:
            rfile_path = os.path.join(regs_folder, file_name)
            os.remove(rfile_path)

        ############### PICKLE ###################
        # COPY PICKLE FILE TO PREPROCESS_FOLDER (Manter por último para garantir que os arquivos estão no storage)
        pickle_folder = os.path.join(preprocessing_folder, 'SPED_REPORTS', 'PICKLES')
        os.makedirs(pickle_folder, exist_ok=True)
        shutil.copy(pfile_path, os.path.join(pickle_folder, pfile_name))


    father_by_reg.clear()
    father_by_reg = None
    son_by_father.clear()
    son_by_father = None
    reg_file_pointers.clear()
    reg_file_pointers = None
    file_pointers.clear()
    file_pointers = None
    report_fieldnames.clear()
    report_fieldnames = None
    compl_info.clear()
    compl_info = None
    report_reg_compl.clear()
    report_reg_compl = None
    count_reg_rows.clear()
    count_reg_rows = None
    summary_by_report.clear()
    summary_by_report = None
    summary_by_field.clear()
    summary_by_field = None
    

    return file_info, logs



def generate_report_fieldnames( sped : str, version : str, competencia : str, chield_level = 3 ):
    if sped == 'ECD':
        version = competencia

    report_fieldnames = {}
    sped_dict = load_sped_dictionary( sped, version, SPED_BY_VERSION[sped] )
    prefix = SPED_BY_VERSION[sped][version]['prefixo']
    report_reg_compl = deepcopy( REPORT_REG_COMPL )
    compl_tables = load_compl_tables()
    report_fieldnames = get_report_fieldnames( sped, prefix, sped_dict, chield_level, report_fieldnames, report_reg_compl, compl_tables )

    for reg in report_fieldnames:
        report_fieldnames[reg] = {'CompetenciaArquivo' : 'CompetenciaArquivo', **report_fieldnames[reg], 'RowId' : 'RowId', 'HashFile' : 'HashFile', 'FileName' : 'FileName'}

    return report_fieldnames





if __name__ == '__main__' :
    # file_path = '/home/eduardo/Documentos/STORAGE/TESTE_TABLES/FILES_RECIEVED/14395513000153-177167160113-20170701-20170731-0-55D8E5B916F9DD1D57295BB536DF55B7E6D5A6ED-SPED-EFD.txt'
    # output_path = '/home/eduardo/Documentos/STORAGE/TESTE_TABLES/REPORT'

    # log = LogWriter( output_path )
    # generate_reports_sped( file_path, output_path, log=log )

    output_path = '/home/eduardo/Downloads'
    extract_fieldnames_v2( output_path )

    # load_compl_tables()
