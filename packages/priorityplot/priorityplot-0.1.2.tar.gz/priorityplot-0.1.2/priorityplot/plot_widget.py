from PyQt6.QtWidgets import QWidget, QVBoxLayout, QPushButton, QTabWidget, QTableWidget, QTableWidgetItem, QLineEdit, QLabel, QHBoxLayout, QMessageBox, QFileDialog
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QClipboard
from .model import Task, calculate_and_sort_tasks
import numpy as np
from openpyxl import Workbook
from datetime import datetime
from PyQt6.QtWidgets import QApplication

class PriorityPlotWidget(QWidget):
    def __init__(self, task_list=None):
        super().__init__()
        self.task_list = task_list if task_list is not None else []
        self.dragging = False
        self.drag_index = None
        self.moved_points = set()  # Track which points have been moved
        self.current_annotation = None  # Track current hover annotation
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()
        self.tabs = QTabWidget()
        self.input_tab = QWidget()
        self.plot_tab = QWidget()
        self.table_tab = QWidget()
        self.tabs.addTab(self.input_tab, "Input Goals")
        self.tabs.addTab(self.plot_tab, "Plot")
        self.tabs.addTab(self.table_tab, "Table")
        self.tabs.setTabEnabled(1, False)
        self.tabs.setTabEnabled(2, False)
        layout.addWidget(self.tabs)
        self.setLayout(layout)
        self.initInputTab()
        self.initPlotTab()
        self.initTableTab()

    def initInputTab(self):
        layout = QVBoxLayout()
        form_layout = QHBoxLayout()
        self.task_input = QLineEdit()
        self.task_input.setPlaceholderText("Task name")
        self.task_input.returnPressed.connect(self.add_task)
        form_layout.addWidget(QLabel("Task:"))
        form_layout.addWidget(self.task_input)
        self.add_button = QPushButton("Add Goal")
        self.add_button.clicked.connect(self.add_task)
        form_layout.addWidget(self.add_button)
        
        # Add clipboard import button
        self.clipboard_button = QPushButton("Add Goals from Clipboard")
        self.clipboard_button.clicked.connect(self.add_goals_from_clipboard)
        form_layout.addWidget(self.clipboard_button)
        
        # Add test goals button
        self.test_button = QPushButton("Add Test Goals")
        self.test_button.clicked.connect(self.add_test_goals)
        form_layout.addWidget(self.test_button)
        
        layout.addLayout(form_layout)
        self.input_table = QTableWidget()
        self.input_table.setColumnCount(1)
        self.input_table.setHorizontalHeaderLabels(["Task"])
        layout.addWidget(self.input_table)
        self.done_button = QPushButton("Done")
        self.done_button.clicked.connect(self.finish_input)
        layout.addWidget(self.done_button)
        self.input_tab.setLayout(layout)
        self.refresh_input_table()

    def add_test_goals(self):
        test_goals = [
            ("Complete Project Proposal", 4.5, 3.0),
            ("Review Code Changes", 3.0, 2.0),
            ("Team Meeting", 2.5, 1.5),
            ("Update Documentation", 3.5, 4.0),
            ("Bug Fixing", 4.0, 2.5),
            ("Client Presentation", 5.0, 4.0),
            ("Code Refactoring", 3.5, 5.0),
            ("Unit Testing", 4.0, 3.0),
            ("Performance Optimization", 4.5, 6.0),
            ("Security Audit", 5.0, 4.5),
            ("Database Migration", 4.0, 7.0),
            ("API Integration", 3.5, 3.5),
            ("User Training", 3.0, 2.0),
            ("System Backup", 2.5, 1.0),
            ("Deployment Planning", 4.0, 2.0),
            ("Code Review", 3.5, 1.5),
            ("Feature Implementation", 4.5, 5.0),
            ("Technical Documentation", 3.0, 4.0),
            ("Bug Triage", 3.5, 2.0),
            ("System Monitoring", 2.5, 1.5)
        ]
        
        for task_name, value, time in test_goals:
            self.task_list.append(Task(task_name, value, time))
        
        self.refresh_input_table()

    def add_task(self):
        task = self.task_input.text().strip()
        if not task:
            QMessageBox.warning(self, "Input Error", "Task name cannot be empty.")
            return
        # Initialize with default values within our valid range
        self.task_list.append(Task(task, 3.0, 4.0))  # Default to middle of our ranges
        self.task_input.clear()
        self.refresh_input_table()

    def add_goals_from_clipboard(self):
        clipboard = QApplication.clipboard()
        text = clipboard.text().strip()
        
        if not text:
            QMessageBox.warning(self, "Clipboard Empty", "No text found in clipboard.")
            return
            
        goals = [goal.strip() for goal in text.split('\n') if goal.strip()]
        
        if not goals:
            QMessageBox.warning(self, "No Valid Goals", "No valid goals found in clipboard text.")
            return
            
        for goal in goals:
            self.task_list.append(Task(goal, 3.0, 4.0))  # Default to middle of our ranges
            
        self.refresh_input_table()
        QMessageBox.information(self, "Success", f"Added {len(goals)} goals from clipboard.")

    def refresh_input_table(self):
        self.input_table.setRowCount(len(self.task_list))
        for i, t in enumerate(self.task_list):
            self.input_table.setItem(i, 0, QTableWidgetItem(t.task))

    def finish_input(self):
        if not self.task_list:
            QMessageBox.warning(self, "No Tasks", "Please add at least one task.")
            return
        self.tabs.setTabEnabled(1, True)
        self.tabs.setTabEnabled(2, True)
        self.tabs.setCurrentWidget(self.plot_tab)
        self.update_plot()

    def initPlotTab(self):
        layout = QVBoxLayout()
        self.figure = Figure(figsize=(5, 4), facecolor='#353535')
        self.canvas = FigureCanvas(self.figure)
        self.ax = self.figure.add_subplot(111)
        
        # Set modern plot style
        self.ax.set_facecolor('#353535')
        self.ax.grid(True, linestyle='--', alpha=0.3, color='#555555')
        self.ax.spines['bottom'].set_color('#555555')
        self.ax.spines['top'].set_color('#555555')
        self.ax.spines['left'].set_color('#555555')
        self.ax.spines['right'].set_color('#555555')
        
        # Set labels with modern styling
        self.ax.set_xlabel('Value', color='white', fontsize=10, fontweight='bold')
        self.ax.set_ylabel('Time (hours)', color='white', fontsize=10, fontweight='bold')
        self.ax.set_title('Task Priority Plot', color='white', fontsize=12, fontweight='bold', pad=20)
        
        # Style the ticks
        self.ax.tick_params(colors='white', which='both')
        
        # Set fixed axis limits
        self.ax.set_xlim(0, 6)
        self.ax.set_ylim(0, 8)
        
        # Connect mouse events
        self.canvas.mpl_connect('button_press_event', self.on_press)
        self.canvas.mpl_connect('button_release_event', self.on_release)
        self.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.canvas.mpl_connect('motion_notify_event', self.on_hover)
        
        self.scatter = self.ax.scatter(
            [t.value for t in self.task_list],
            [t.time for t in self.task_list],
            picker=True,
            alpha=0.7
        )
        
        # Adjust figure layout
        self.figure.tight_layout()
        self.canvas.draw()
        
        layout.addWidget(self.canvas)
        self.apply_button = QPushButton('Apply')
        self.apply_button.clicked.connect(self.showTable)
        layout.addWidget(self.apply_button)
        self.plot_tab.setLayout(layout)

    def on_press(self, event):
        if event.inaxes != self.ax:
            return
        contains, ind = self.scatter.contains(event)
        if contains:
            self.dragging = True
            self.drag_index = ind["ind"][0]

    def on_motion(self, event):
        if not self.dragging or self.drag_index is None or event.inaxes != self.ax:
            return
        
        # Update task values
        self.task_list[self.drag_index].value = event.xdata
        self.task_list[self.drag_index].time = event.ydata
        self.moved_points.add(self.drag_index)  # Mark this point as moved
        
        # Update scatter plot data directly for smooth movement
        x_data = [t.value for t in self.task_list]
        y_data = [t.time for t in self.task_list]
        self.scatter.set_offsets(np.column_stack([x_data, y_data]))
        
        # Update the annotation position
        for i, annotation in enumerate(self.ax.texts):
            if i == self.drag_index:
                annotation.set_position((event.xdata, event.ydata))
        
        self.canvas.draw_idle()  # Use draw_idle for smoother updates

    def on_release(self, event):
        if self.dragging:
            # Do a full redraw when drag is complete
            self.update_plot()
        self.dragging = False
        self.drag_index = None

    def on_hover(self, event):
        if event.inaxes != self.ax:
            if self.current_annotation:
                self.current_annotation.set_visible(False)
                self.current_annotation = None
                self.canvas.draw_idle()
            return

        contains, ind = self.scatter.contains(event)
        if contains:
            pos = ind["ind"][0]
            task = self.task_list[pos]
            
            # Remove previous annotation if it exists
            if self.current_annotation:
                self.current_annotation.set_visible(False)
            
            # Create new annotation with task name and values
            text = f"{task.task}\nValue: {task.value:.1f}\nTime: {task.time:.1f}"
            self.current_annotation = self.ax.annotate(
                text,
                xy=(task.value, task.time),
                xytext=(10, 10),
                textcoords='offset points',
                bbox=dict(
                    boxstyle='round,pad=0.5',
                    fc='#2a82da',
                    ec='#555555',
                    alpha=0.9
                ),
                color='white',
                fontsize=9,
                fontweight='bold',
                arrowprops=dict(
                    arrowstyle='->',
                    connectionstyle='arc3,rad=0.2',
                    color='#555555',
                    linewidth=1.5
                )
            )
            self.canvas.draw_idle()
        elif self.current_annotation:
            self.current_annotation.set_visible(False)
            self.current_annotation = None
            self.canvas.draw_idle()

    def update_plot(self):
        self.ax.clear()
        
        # Reapply modern styling
        self.ax.set_facecolor('#353535')
        self.ax.grid(True, linestyle='--', alpha=0.3, color='#555555')
        self.ax.spines['bottom'].set_color('#555555')
        self.ax.spines['top'].set_color('#555555')
        self.ax.spines['left'].set_color('#555555')
        self.ax.spines['right'].set_color('#555555')
        
        # Set labels with modern styling
        self.ax.set_xlabel('Value', color='white', fontsize=10, fontweight='bold')
        self.ax.set_ylabel('Time (hours)', color='white', fontsize=10, fontweight='bold')
        self.ax.set_title('Task Priority Plot', color='white', fontsize=12, fontweight='bold', pad=20)
        
        # Style the ticks
        self.ax.tick_params(colors='white', which='both')
        
        # Maintain fixed axis limits
        self.ax.set_xlim(0, 6)
        self.ax.set_ylim(0, 8)
        
        # Create arrays for all points
        x_data = [t.value for t in self.task_list]
        y_data = [t.time for t in self.task_list]
        
        # Calculate scores and find the top 3 tasks
        for task in self.task_list:
            task.calculate_score()
        
        # Get indices sorted by score (highest first)
        sorted_indices = sorted(range(len(self.task_list)), key=lambda i: self.task_list[i].score, reverse=True)
        top_3_indices = sorted_indices[:3]  # Get top 3 tasks
        
        # Create color array based on whether points have been moved
        colors = ['#2a82da' if i in self.moved_points else '#e74c3c' for i in range(len(self.task_list))]
        
        # Create scatter plot for all points except the top 3
        non_top_indices = [i for i in range(len(self.task_list)) if i not in top_3_indices]
        if non_top_indices:
            self.ax.scatter(
                [x_data[i] for i in non_top_indices],
                [y_data[i] for i in non_top_indices],
                c=[colors[i] for i in non_top_indices],
                picker=True,
                alpha=0.7,
                s=100
            )
        
        # Plot the top 3 tasks with circled numbers
        for rank, task_index in enumerate(top_3_indices, 1):
            if task_index < len(self.task_list):  # Safety check
                task_x = x_data[task_index]
                task_y = y_data[task_index]
                self.ax.plot(task_x, task_y, 'o', markersize=20, markerfacecolor='none', 
                            markeredgecolor=colors[task_index], markeredgewidth=2)
                self.ax.text(task_x, task_y, str(rank), ha='center', va='center', 
                            fontsize=14, fontweight='bold', color=colors[task_index])
        
        # Update the scatter reference for event handling
        self.scatter = self.ax.scatter(x_data, y_data, c=colors, picker=True, alpha=0)
        
        # Adjust figure layout
        self.figure.tight_layout()
        self.canvas.draw()

    def initTableTab(self):
        layout = QVBoxLayout()
        self.table = QTableWidget()
        layout.addWidget(self.table)
        
        # Add export button
        self.export_button = QPushButton('Export to Excel')
        self.export_button.clicked.connect(self.export_to_excel)
        layout.addWidget(self.export_button)
        
        self.table_tab.setLayout(layout)

    def export_to_excel(self):
        if not self.task_list:
            QMessageBox.warning(self, "No Data", "There are no tasks to export.")
            return
            
        # Get save file path
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            f"priority_plot_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Task Priorities"
            
            # Add headers
            headers = ['Task', 'Value', 'Time (hours)', 'Priority Score']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            sorted_tasks = calculate_and_sort_tasks(self.task_list)
            for row, task in enumerate(sorted_tasks, 2):
                ws.cell(row=row, column=1, value=task.task)
                ws.cell(row=row, column=2, value=task.value)
                ws.cell(row=row, column=3, value=task.time)
                ws.cell(row=row, column=4, value=task.score)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(file_path)
            QMessageBox.information(self, "Success", "Data exported successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export data: {str(e)}")

    def showTable(self):
        sorted_tasks = calculate_and_sort_tasks(self.task_list)
        self.table.setRowCount(len(sorted_tasks))
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['Task', 'Value', 'Time', 'Priority Score'])
        for i, t in enumerate(sorted_tasks):
            self.table.setItem(i, 0, QTableWidgetItem(t.task))
            self.table.setItem(i, 1, QTableWidgetItem(f"{t.value:.2f}"))
            self.table.setItem(i, 2, QTableWidgetItem(f"{t.time:.2f}"))
            self.table.setItem(i, 3, QTableWidgetItem(f"{t.score:.2f}"))
        self.tabs.setCurrentWidget(self.table_tab) 