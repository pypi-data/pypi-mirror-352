"""
Functions to transform data from Excel.
Main function being:

process_fout_sheets
"""
import datetime
import re
import logging
from collections import namedtuple
from openpyxl.workbook.workbook import Workbook
import pandas as pd
from dqchecks.exceptions import (
    EmptyRowsPatternCheckError,
    ColumnHeaderValidationError,)

logging.basicConfig(level=logging.INFO)

# Define namedtuple for context
ProcessingContext = namedtuple(
    'ProcessingContext', ['org_cd', 'submission_period_cd', 'process_cd',
                          'template_version', 'last_modified']
)

def is_valid_regex(pattern: str) -> bool:
    """
    Check if a given string is a valid regex
    """
    try:
        re.compile(pattern)  # Try to compile the regex pattern
        return True  # If no exception, it's a valid regex
    except re.error:  # If an exception is raised, it's not a valid regex
        return False

def validate_workbook(wb: Workbook):
    """
    Validates if the provided workbook is an instance of openpyxl Workbook.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("The 'wb' argument must be a valid openpyxl workbook object.")

def validate_context(context: ProcessingContext):
    """
    Validates the fields in the context object.
    """
    if not isinstance(context.org_cd, str) or not context.org_cd:
        raise ValueError("The 'org_cd' argument must be a non-empty string.")

    if not isinstance(context.submission_period_cd, str) or not context.submission_period_cd:
        raise ValueError("The 'submission_period_cd' argument must be a non-empty string.")

    if not isinstance(context.process_cd, str) or not context.process_cd:
        raise ValueError("The 'process_cd' argument must be a non-empty string.")

    if not isinstance(context.template_version, str) or not context.template_version:
        raise ValueError("The 'template_version' argument must be a non-empty string.")

    if not isinstance(context.last_modified, datetime.datetime):
        raise ValueError("The 'last_modified' argument must be a datetime object.")

def validate_observation_patterns(observation_patterns: list[str]):
    """
    Validates the observation patterns argument.
    """
    if not isinstance(observation_patterns, list)\
            or not all(isinstance(i, str) for i in observation_patterns)\
            or not all(is_valid_regex(i) for i in observation_patterns):
        raise ValueError("The 'observation_patterns' argument needs to be a list of regex strings.")

def extract_fout_sheets(wb: Workbook, fout_patterns: list[str]):
    """
    Extracts sheets from the workbook whose names match any of the given regex patterns.
    
    Args:
        wb (Workbook): The Excel workbook object.
        fout_patterns (list[str]): A list of regex patterns to match sheet names.
    
    Returns:
        List[str]: A list of matching sheet names.

    Raises:
        ValueError: If no matching sheets are found.
    """
    regexes = [re.compile(p) for p in fout_patterns]

    matching_sheets = [
        sheet for sheet in wb.sheetnames
        if any(regex.match(sheet) for regex in regexes)
    ]

    if not matching_sheets:
        raise ValueError(
            f"No sheets matching patterns {fout_patterns} found. Available sheets: {wb.sheetnames}"
        )

    return matching_sheets


def read_sheets_data(wb: Workbook, fout_sheets: list):
    """
    Reads data from the sheets into pandas DataFrames.
    """
    df_list = []
    for sheetname in fout_sheets:
        data = wb[sheetname].iter_rows(min_row=2, values_only=True)
        try:
            headers = next(data)
        except StopIteration as exc:
            raise ValueError(f"Sheet '{sheetname}' is empty or has no data.") from exc

        df = pd.DataFrame(data, columns=headers)
        df["Sheet_Cd"] = sheetname  # Add a column for the sheet name
        df_list.append(df)
    return df_list

def clean_data(df_list: list):
    """
    Drops rows with NaN values and checks if any dataframe is empty.
    """
    df_list = [
        df.dropna(how='all', subset=df.columns.difference(['Sheet_Cd']))
        for df in df_list
    ]
    if any(i.empty for i in df_list):
        raise ValueError("No valid data found after removing rows with NaN values.")
    return df_list

def process_observation_columns(df: pd.DataFrame, observation_patterns: list[str]):
    """
    Identifies and returns the observation period columns based on the provided patterns.
    """
    observation_period_columns = []
    for observation_pattern in observation_patterns:
        observation_period_columns += list(df.filter(regex=observation_pattern).columns.tolist())
    return set(observation_period_columns)

def process_df(df: pd.DataFrame, context: ProcessingContext, observation_patterns: list[str]):
    """
    Processes a single dataframe by melting it and adding context columns.
    """
    observation_period_columns = process_observation_columns(df, observation_patterns)
    if not observation_period_columns:
        raise ValueError("No observation period columns found in the data.")

    # Get the ID columns (all columns except observation period columns)
    id_columns = set(df.columns.tolist()) - observation_period_columns

    # Pivot the DataFrame to melt observation period columns into rows
    pivoted_df = df.melt(
        id_vars=id_columns,
        var_name="Observation_Period_Cd",
        value_name="Measure_Value"
    )

    # Add static context columns to the pivoted DataFrame
    pivoted_df["Organisation_Cd"] = context.org_cd
    pivoted_df["Submission_Period_Cd"] = context.submission_period_cd
    pivoted_df["Process_Cd"] = context.process_cd
    pivoted_df["Template_Version"] = context.template_version
    pivoted_df["Submission_Date"] = context.last_modified  # Use the last modified date
    pivoted_df["Section_Cd"] = "--placeholder--"
    pivoted_df["Cell_Cd"] = "--placeholder--"

    # Convert all columns to strings for consistency
    pivoted_df = pivoted_df.astype(str)

    # Define a mapping for renaming columns to the final desired format
    column_rename_map = {
        'Organisation_Cd': 'Organisation_Cd',
        'Submission_Period_Cd': 'Submission_Period_Cd',
        'Observation_Period_Cd': 'Observation_Period_Cd',
        'Process_Cd': 'Process_Cd',
        'Template_Version': 'Template_Version',
        'Sheet_Cd': 'Sheet_Cd',
        'Reference': 'Measure_Cd',
        'Measure_Value': 'Measure_Value',
        'Item description': 'Measure_Desc',
        'Unit': 'Measure_Unit',
        'Model': 'Model_Cd',
        'Submission_Date': 'Submission_Date',
        "Section_Cd": "Section_Cd",
        "Cell_Cd": "Cell_Cd",
    }

    # Rename the columns according to the mapping
    pivoted_df = pivoted_df.rename(columns=column_rename_map)

    # Reorder the columns to match the desired output format
    ordered_columns = list(column_rename_map.values())
    pivoted_df = pivoted_df[ordered_columns]

    return pivoted_df

def check_empty_rows(wb: Workbook, sheet_names: list[str]):
    # pylint: disable=C0301
    """
    Validates that specified sheets in a workbook contain only empty cells in specific rows.

    This function performs two checks on each provided worksheet:
      1. Verifies that all cells in row 3 (under the header) are empty.
      2. Verifies that all cells in row 1 (the top row), excluding the third column, are empty.

    If any sheet fails either check, a custom EmptyRowsPatternCheckError is raised, indicating which sheets failed.

    Parameters:
        wb (Workbook): An openpyxl Workbook instance containing the sheets to check.
        sheet_names (list[str]): A list of worksheet names to validate.

    Returns:
        bool: True if all sheets pass the checks.

    Raises:
        TypeError: If 'wb' is not a Workbook or 'sheet_names' is not a list of strings.
        ValueError: If 'sheet_names' is empty or contains names not found in the workbook.
        EmptyRowsPatternCheckError: If any sheet contains non-empty values in the checked rows.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("Expected an openpyxl Workbook instance for 'wb'.")
    if not isinstance(sheet_names, list) or not all(isinstance(name, str) for name in sheet_names):
        raise TypeError("Expected 'sheet_names' to be a list of strings.")
    if not sheet_names:
        raise ValueError("'sheet_names' list cannot be empty.")

    under_header_bad_sheet_names = []
    top_row_bad_sheet_names = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]

        # Check under header row (row 3)
        under_header_row = sheet.iter_rows(min_row=3, values_only=True)
        under_header_row_vals = list(next(under_header_row, []))
        if set(under_header_row_vals) not in [{None}, {"", None}, {""}]:
            under_header_bad_sheet_names.append(sheet_name)

        # Check top row (row 1), with 3rd element removed
        top_row = sheet.iter_rows(min_row=1, values_only=True)
        top_row_vals = list(next(top_row, []))
        if len(top_row_vals) > 2:
            del top_row_vals[2]
        if set(top_row_vals) not in [{None}, {"", None}, {""}]:
            top_row_bad_sheet_names.append(sheet_name)

    if under_header_bad_sheet_names or top_row_bad_sheet_names:
        raise EmptyRowsPatternCheckError(under_header_bad_sheet_names, top_row_bad_sheet_names)

    return True  # Validation passed

def check_column_headers(wb: Workbook, sheet_names: list[str]):
    """
    Validates that each sheet has the required columns in the correct order starting from row 2.

    Args:
        wb (Workbook): The openpyxl workbook object.
        sheet_names (list[str]): List of sheet names to check.

    Raises:
        TypeError: If wb is not a Workbook, or sheet_names is not a list of strings.
        ValueError: If sheet_names is empty or contains names not in the workbook.
        ColumnHeaderValidationError: If any sheet has missing or misordered expected columns.

    Returns:
        True: If all sheets pass the header validation.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("Expected an openpyxl Workbook instance for 'wb'.")
    if not isinstance(sheet_names, list) or not all(isinstance(name, str) for name in sheet_names):
        raise TypeError("Expected 'sheet_names' to be a list of strings.")
    if not sheet_names:
        raise ValueError("'sheet_names' list cannot be empty.")
    if not all(name in wb.sheetnames for name in sheet_names):
        raise ValueError("One or more sheet names are not present in the workbook.")

    expected_columns = ["Acronym", "Reference", "Item description", "Unit", "Model"]
    bad_sheets = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        header_rows = sheet.iter_rows(min_row=2, max_row=2, values_only=True)
        header = next(header_rows, ())

        # Keep only the expected columns in the order they appear in the sheet
        filtered_header = [col for col in header if col in expected_columns]

        if filtered_header != expected_columns:
            bad_sheets.append(sheet_name)

    if bad_sheets:
        raise ColumnHeaderValidationError(bad_sheets, expected_columns)

    return True


def process_fout_sheets(
        wb: Workbook,
        context: ProcessingContext,
        observation_patterns: list[str],
        fout_patterns: list[str],
        ):
    """
    Processes all sheets in the given Excel workbook that start with 'fOut_' and extracts 
    data into a pandas DataFrame. Each sheet is expected to have columns representing 
    observations for various periods, and the function performs a transformation to normalize 
    the data into a consistent format.

    Args:
        wb (openpyxl.workbook.workbook.Workbook): The openpyxl Workbook object
            containing the data to process.
        context (ProcessingContext): The context object containing the organization code, 
                                     submission period code, process code, template version, 
                                     and last modified timestamp.
        observation_patterns (list[str]): A list of regular expression patterns
            used to match columns representing observation periods in the data.
        fout_patterns (list[str]): A list of regex patterns to match sheet names.

    Returns:
        pd.DataFrame: A pandas DataFrame containing the processed data from all matching sheets, 
                      with columns like 'Organisation_Cd', 'Observation_Period_Cd', 'Measure_Value', 
                      and other related columns.

    Raises:
        ValueError: If any of the input values are invalid, if no matching sheets are found, 
                    if no observation period columns are found, or if no valid data is found after 
                    dropping NaN rows.
        TypeError: If the provided workbook is not of the correct type
            (i.e., not an openpyxl Workbook).
        Exception: For any other unexpected errors.

    Example:
        # Assuming 'wb' is a valid openpyxl Workbook and 'context' 
        # is a valid ProcessingContext object
        wb = openpyxl.load_workbook("data.xlsx")
        context = ProcessingContext(
            org_cd="ORG001",
            submission_period_cd="2025Q1",
            process_cd="PROCESS01",
            template_version="v1.0",
            last_modified=datetime.datetime(2025, 3, 4, 10, 30)
        )
        # Regex pattern to match observation period columns
        observation_patterns = [r"^ObsPeriod_\\d+$"]

        # Call the function
        processed_data = process_fout_sheets(wb, context, observation_patterns)

        # The result will be a DataFrame containing the processed data
        print(processed_data.head())  # Prints the first few rows of the processed DataFrame
    """

    # Validate inputs
    validate_workbook(wb)
    validate_context(context)
    validate_observation_patterns(observation_patterns)

    # Check if the workbook is in data_only mode (warn if not)
    if not wb.data_only:
        logging.warning("Reading in non data_only mode. Some data may not be accessible.")

    logging.info("Using observation patterns: %s", observation_patterns)

    # Extract sheets that start with 'fOut_'
    fout_sheets = extract_fout_sheets(wb, fout_patterns)

    # Check if fOut tabs follow expected pattern with empty rows
    assert check_empty_rows(wb, fout_sheets)

    # Check that columns exist and in correct order
    assert check_column_headers(wb, fout_sheets)

    # Read data from the sheets
    df_list = read_sheets_data(wb, fout_sheets)

    # Clean data by dropping rows with NaN values and checking for empty DataFrames
    df_list = clean_data(df_list)

    # Process each DataFrame and collect the results
    melted_dfs = [process_df(df, context, observation_patterns) for df in df_list]

    # Concatenate all the melted DataFrames into one final DataFrame
    final_df = pd.concat(melted_dfs, ignore_index=True)

    return final_df
